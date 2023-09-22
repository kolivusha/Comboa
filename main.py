from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter.messagebox import showinfo
import os
import time
from datetime import datetime, date, timedelta
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import pandas as pd
import csv
from configparser import ConfigParser
import clipboard
# from gtts import gTTS
from playsound import playsound
import threading
import win32com.client
# from pywinauto.application import Application
import win32gui
import pygetwindow as gw
import sys
import numpy as np
from natsort import natsorted
from sqlalchemy import create_engine,text


# import matplotlib
# import matplotlib.backends.backend_wx
# matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg, NavigationToolbar2Tk)

from matplotlib.backend_bases import key_press_handler
from matplotlib.figure import Figure

# from pydub import AudioSegment #need ffmpeg files to run, and add pause to mp3s
# pyinstaller main.py --onefile --noconsole -n Comboa --hidden-import pandas.plotting._matplotlib
# --uac-admin
# pyinstaller main.py --onefile --noconsole -n ComboaU --hidden-import pandas.plotting._matplotlib

config = ConfigParser()
config.read("config.ini", encoding='utf-8')
prefixdict = {i: config['Test Centers'][i].split(';') for i in config.options('Test Centers')}
userpath = os.environ['USERPROFILE']


def listFilesInfolder(path):
    files = []
    for i in os.listdir(path):
        files.append(i)
    # print(len(files))
    return files


def getfileLSfromdate(runname):
    Folder1 = os.path.join(userpath + config['settings']['Pathfolder1'])
    folder1ls = (listFilesInfolder(Folder1))
    todaysfilesLS = []
    for filename in folder1ls:
        if filename.startswith(f'{runname}'[:7]):
            todaysfilesLS.append(filename[:-5])
    return todaysfilesLS


def GenerateRunname():
    global todaysfilesLS, userpath, runname
    # today = date.today().strftime("%d%m%Y")
    today = (datetime.now() - timedelta(hours=5)).strftime("%d%m%Y")  # uses todays date until 5am tommorow
    region = config['settings']['Region']
    if expvar.get() == 'PCR':
        Folder1 = os.path.join(userpath + config['settings']['Pathfolder1'])
        name = 'Run'
    elif expvar.get() == 'ELISA':
        Folder1 = os.path.join(userpath + config['settings']['ElISAsaves'])
        name = expTypeVar.get()
    print(f"{name=}")
    suggestedName = f"{today}_{region}_{name}1"
    folder1ls = (listFilesInfolder(Folder1))
    todaysfilesLS, numbersls = [], []
    for filename in folder1ls:
        # try:
        #     print (runname)
        # except:
        #     print('runname is empty')
        if filename.startswith(today) and filename[-6].isdigit():
            todaysfilesLS.append(filename[:-5])  # removes ".xlsx" from the name
            # suggestedName = f'{today}_Run{len(todaysfilesLS)+1}' #simple bias on number of todays files, can be more complicated, knowing exact runnumber
    # print(todaysfilesLS)
    if len(todaysfilesLS) > 0:
        for i in todaysfilesLS:
            try:
                numbersls.append(int(i.split(f"{today}_{region}_{name}")[1]))
            except:
                print(f'Cannot convert runname: {i} to integer')
        # print(f'{numbersls=}')
        if len(numbersls) == 1:
            suggestedName = f"{today}_{region}_{name}2"
        else:
            try:
                suggestedName = f"{today}_{region}_{name}{max(numbersls) + 1}"
            except ValueError as e:
                suggestedName = f"{today}_{region}_{name}1"
                print(f"{e=}")
                print('Exception tiggered in generating runname')
    print(f'{suggestedName=}')
    return suggestedName


def scanforresolution(fullscan=1):
    if expvar.get() == 'ELISA':
        return
    # root.config(cursor="wait")
    t = time.perf_counter()
    global runname, dfResolution, userpath, yesterdayreps, dfDaily, exportreportls
    print(f"{yesterdayreps.get()=}")
    exportreportls = []

    def getfileLSfromdate():
        Folder1 = os.path.join(userpath + config['settings']['Pathfolder1'])
        folder1ls = (listFilesInfolder(Folder1))
        todaysfilesLS = []
        if yesterdayreps.get() == TRUE:
            yestraday = (datetime.strptime(f'{runname}'[0:8], "%d%m%Y") - timedelta(days=1)).strftime("%d%m%Y")
            print(f"{yestraday=}")
            for filename in folder1ls:
                if filename.startswith(f'{runname}'[0:8]) or filename.startswith(yestraday):
                    todaysfilesLS.append(filename)
        else:
            for filename in folder1ls:
                if filename.startswith(f'{runname}'[0:8]):
                    todaysfilesLS.append(filename)
        return todaysfilesLS

    def assignresult(row):
        if row['Manual Result'] == 'None':
            return row['Result Auto']
        else:
            return row['Manual Result']

    def ispoool(row):
        if len(row['IDs']) > 1:
            return 1
        else:
            return 0

    def toResolve(row):
        if row['w.number'] in [95, 96]:
            return 0
        if row['ispool'] == 1 and row['result'] == 'Positive':
            return 1
        if row['result'] == 'Repeat':
            return 1
        else:
            return 0

    def isResolved(row):  # check if ID is mentioned second time, in non-poll
        id = row['ID']
        a = df.loc[(df['ID'] == id)
                   & (df['ispool'] == 0)
                   # & (df['result']=='Unknown')
                   & (df['runwell'] != row['runwell'])]
        if a.empty:
            return 'Not Found'
        else:
            return a[['runwell']].values[0][0]

    df = pd.DataFrame()

    def convertToInt(row):
        a = str(row['ID'])
        if a.endswith('.0'):
            return a[:-2]
        else:
            return a

    todaysfilesLS = getfileLSfromdate()
    for file in todaysfilesLS:
        try:
            if file.startswith('~$'):
                continue
            Folder1 = os.path.join(userpath + config['settings']['Pathfolder1'])
            print(Folder1 + '\\' + file)
            wb = load_workbook(filename=Folder1 + '\\' + file)
            ws = wb.active
            numberoflayers = ws['G2'].value
            run = pd.read_excel(Folder1 + '\\' + file, skiprows=3, )
            layernames = [f'layer {x + 1}' for x in range(numberoflayers)]
            run = run.astype(object).replace(np.nan, 'None')
            run['ID'] = run[layernames].dropna().values.tolist()
            run['ID'] = [[x for x in inner_list if x != 'None'] for inner_list in run['ID']]
            run['IDs'] = run['ID']
            run = run.drop(run[run['ID'].map(len) < 1].index)  ### remove empty wells
            run['daterun'] = file[:-5]
            run['runwell'] = run['daterun'].astype(str) + ' ' + run['w.number'].astype(str)
            run['date'] = datetime.strptime(file[:8], '%d%m%Y').date()
            run['week'] = datetime.strptime(file[:8], '%d%m%Y').isocalendar().week
            run['result'] = run.apply(assignresult, axis=1)
            run['ispool'] = run.apply(ispoool, axis=1)
            run['toResolve'] = run.apply(toResolve, axis=1)
            run.drop(layernames, axis=1, inplace=True)
            run = run.explode('ID')
            run['ID'] = run.apply(convertToInt, axis=1)
            # print(run['ID'])
            df = pd.concat([df, run])
        except:
            print('skipped')
            print(sys.exc_info())
        try:
            exportreportls.append(f"{file[:-5]}#{(ws['k2'].value)[-19:]}\n")
        except:
            exportreportls.append(f"{file[:-5]}#not exported\n")
    dfDaily = df
    if df.empty:
        dfResolution = pd.DataFrame
        print('df is empty')
        resolutionbutton()
        # root.config(cursor="")
        return
    dfResolution = df.loc[df['toResolve'] == 1]
    if dfResolution.empty:
        dfResolution = pd.DataFrame
        print('dfResolution is empty')
        resolutionbutton()
        # root.config(cursor="")
        return
    dfResolution['isResolved'] = dfResolution.apply(isResolved, axis=1)
    # pd.set_option('display.max_columns', 16)
    # print(dfResolution[['runwell',
    #                     # 'w.name','ID','result',
    #                     'isResolved','ID']])
    # print(dfResolution)
    try:
        print(f'to repeat {len(dfResolution.index) - dfResolution.isResolved.value_counts()["Not Found"]}'
              f'/{len(dfResolution.index)}')
    except:
        pass
    resolutionbutton()
    elapsed_time = time.perf_counter() - t
    # root.config(cursor="")
    print(f"scanforresolution {elapsed_time=}")


def raise_frame(frame):
    frame.tkraise()


def on_closing():
    if messagebox.askokcancel("Quit", "Do you want to quit?"):
        root.quit()


root = Tk()
style = ttk.Style()
version = '1.10 ALLCHECK'
Folder0 = os.path.join(userpath + config['settings']['Pathfolder0'])
root.tk.call('source', f'{Folder0}//audio//defaults.tcl')
style.theme_use('default')

expvar = tk.StringVar()
expvar.set('PCR')
expTypeVar = tk.StringVar()
expTypeVar.set('SarsAB')
runname = GenerateRunname()
totalNumberofLayers = int(config['settings']['Layers'])
root.title(f'{runname}-Comboa {version}')
root.resizable(False, False)
# root.geometry('1835x1009+1200+0')
root.geometry('1840x1009+20+20')
poeple = config['settings']['Lab'].split(",")
sup = config['settings']['Sup'].split(",")
doctors = config['settings']['Doctor'].split(",")
global layerSelected, newcell, oldcell, name1t, name2t, name3t, BatchPCRt, BatchRNAt, foundLS, Showposintable, positivetableLS \
    , dfResolution, sampleindicatorls, lang, Editrun, todaysfilesLS, ELISA_results
name1t, name2t, name3t, BatchPCRt, BatchRNAt = None, None, doctors[0], None, None,
Showposintable = 0
layerSelected = 0
Usedwells = 96
verticallist = [0, 12, 24, 36, 48, 60, 72, 84, 1, 13, 25, 37, 49, 61, 73, 85, 2, 14, 26, 38, 50, 62, 74, 86, 3, 15, 27,
                39, 51, 63, 75, 87, 4, 16, 28, 40, 52, 64, 76, 88, 5, 17, 29, 41, 53, 65, 77, 89, 6, 18, 30, 42, 54, 66,
                78, 90, 7, 19, 31, 43, 55, 67, 79, 91, 8, 20, 32, 44, 56, 68, 80, 92, 9, 21, 33, 45, 57, 69, 81, 93, 10,
                22, 34, 46, 58, 70, 82, 94, 11, 23, 35, 47, 59, 71, 83, 95]

listOfLayers = []


def AddLayers(desirednumber):
    global listOfLayers
    listOfLayers = []
    None94ls = [''] * 94
    None96ls = [''] * 96
    None94ls = None94ls + ['NEGATIVE', 'POSITIVE', ]
    if expvar.get() == 'ELISA':
        None94ls = [''] * 96
        print(f"{listOfLayers=}")
    listOfLayers.append(None94ls.copy())
    for i in range(desirednumber - 1):
        listOfLayers.append(None96ls.copy())


AddLayers(totalNumberofLayers)
#
# layer1,layer2,layer3,layer4,layer5=None94ls.copy(),None94ls.copy(),None94ls.copy(),None94ls.copy(),None94ls.copy()
# listOfLayers=[layer1,layer2,layer3,layer4,layer5]

f1 = Frame(root)
f2 = Frame(root)
tree = ttk.Treeview(f2, columns=None, show='headings', height=38)
f3 = Frame(root)
f4 = Frame(root)
plateframe = Frame(f1, bg='coral', width=0, height=0, padx=5, pady=5, )
plateframe.grid(row=0, column=4, sticky="news", rowspan=60)
entryframe = Frame(f1, bg='lavender', padx=220, pady=20, )
entryframe.grid(row=61, column=4, sticky="nes", rowspan=15)
cellview = Frame(f1, bg='honeydew2', padx=0, pady=0, )
cellview.grid(row=52, column=0, sticky="nws", rowspan=30, columnspan=4)

namesfield = Frame(f1, bg='gray95', padx=0, pady=0, )
namesfield.grid(row=0, column=0, sticky="news", rowspan=7, columnspan=4)
controlsfield = Frame(f1, bg='gray95', padx=0, pady=0, )
controlsfield.grid(row=7, column=0, sticky="news", rowspan=18, columnspan=2)
prefixcountfield = Frame(f1, bg='gray95', padx=0, pady=0, )
# prefixcountfield = Frame(f1, bg='magenta', padx=0,pady=7,)
prefixcountfield.grid(row=20, column=0, sticky="news", rowspan=21, columnspan=4)
notesfield = Frame(f1, bg='paleturquoise1', padx=0, pady=0, )
notesfield.grid(row=42, column=0, sticky="news", rowspan=10, columnspan=4)
for frame in (f1, f2, f3, f4):
    frame.grid(row=0, column=0, sticky='news')
Label(f2, text='FRAME 2').grid(row=0, column=0, sticky='E')
Button(f2, text='Go to page 1', command=lambda: [raise_frame(f1), UpdateSelectedcell()]).grid(row=0, column=3,
                                                                                              sticky='w', padx=10,
                                                                                              pady=10)
Button(f3, text='Go to page 1', command=lambda: [raise_frame(f1), UpdateSelectedcell()]).grid(row=0, column=3,
                                                                                              sticky='w', padx=10,
                                                                                              pady=10)
Button(f4, text='Go to page 1', command=lambda: [raise_frame(f1), UpdateSelectedcell()]).grid(row=0, column=3,
                                                                                              sticky='w', padx=10,
                                                                                              pady=10)


### Frame 1###
def id_entrychange(event):
    # global runname
    # runname = id_entry.get()
    # root.title(f'{runname}-Comboa')
    # print(runname)
    global runname
    root.title(f'{id_entry.get()}-Comboa')
    if runname != id_entry.get():
        runname = id_entry.get()
        try:
            LoadFileOne()  # the problem with this is that no way to load file with the same name as starting name
        except ValueError:
            tk.messagebox.showerror(title='File not found', message='File with such Runname cannot be found')
            print(ValueError)
        root.title(f'{runname}-Comboa')
        print(runname)


id_entry = ttk.Label(namesfield, text='Experiment-ID:').grid(row=0, column=0, sticky='E')


def Todaysfilesupdater():
    global todaysfilesLS
    id_entry['value'] = natsorted(todaysfilesLS)[::-1]


id_entry = ttk.Combobox(namesfield, textvariable='Experiment-ID:', value=todaysfilesLS, postcommand=Todaysfilesupdater,
                        font='Sans 11 bold')  # can be change to Entry or Combobox
id_entry.insert(-1, runname)
id_entry.grid(row=0, column=1, columnspan=2, sticky='EW')
id_entry.bind('<Return>', id_entrychange)
id_entry.bind('<FocusOut>', id_entrychange)
id_entry.bind('<<ComboboxSelected>>', id_entrychange)
id_entry.focus()
newrunbutton = tk.Button(namesfield, text=f'New Run', font='Sans 8 bold',
                         command=lambda: [Createnewrun()]).grid(row=2, column=2, sticky='EW', padx=0, pady=0)


# loadnextbutton = tk.Button(namesfield, text=f'Load Next', font='Sans 9 bold',
#                          command=lambda:[Createnewrun()]).grid(row=1, column=2, sticky='EW',padx=0,pady=0)
# loadpreviousbutton = tk.Button(namesfield, text=f'Load Prev', font='Sans 9 bold',
#                          command=lambda:[Createnewrun()]).grid(row=2, column=2, sticky='EW',padx=0,pady=0)
def Createnewrun():
    # SaveFileOne()
    global layerSelected, newcell, listOfLayers, runname, todaysfilesLS, pcr_resultsLS, sampleindicatorls, totalsamplesOnlyls, \
        positivetableLS, tree, dfFAM, exporttime
    dfFAM = pd.DataFrame
    listOfLayers = []
    totalsamplesOnlyls = []
    positivetableLS = []
    pcr_resultsLS = []
    AddLayers(totalNumberofLayers)
    runname = GenerateRunname()
    sampleindicatorlsmaker()
    Genarate_buttons_fornewlayers()
    exporttime.set(f'')
    print(f'{runname}')
    # id_entrychange(None)
    notesentry.insert(1.0, 'Notes:')
    layerSelected = 0
    newcell = 0
    thread = threading.Thread(None, scanforresolution, daemon=True)
    thread.start()
    UpdateLayer()
    UpdateSelectedcell()
    root.title(f'{runname} -Comboa')
    # print(f'{todaysfilesLS }')
    # id_entry = ttk.Combobox(namesfield, textvariable='Experiment-ID:', value=todaysfilesLS)
    id_entry.delete(0, tk.END)
    id_entry.insert(0, runname)
    notesentry.delete(1.0, tk.END)
    notesentry.insert(1.0, 'Notes:')
    sample_entryfield.delete("0", tk.END)
    # SaveFileOne()
    tree.destroy()


def startRun():
    print(id_entry.get())
    thread = threading.Thread(None, scanforresolution, daemon=True)
    thread.start()
    #### get all fields, wokring file###
    print('Starting run')


# startRunbutton = ttk.Checkbutton(f1, text=f'Start Run', font='Sans 10 bold', command=startRun, bd=2).grid(row=6,column=0,sticky='E', )
yesterdayreps = tk.BooleanVar(value=False)
yesterdayrepbttn = ttk.Checkbutton(controlsfield, text='Include yesterday', command=scanforresolution,
                                   variable=yesterdayreps,
                                   onvalue=1,
                                   offvalue=0,
                                   )
yesterdayrepbttn.grid(row=1, column=1, sticky='W', )


def name1change(event):
    global name1t
    name1_cb.select_clear()
    name1t = name1.get()
    print(name1t)


def name2change(event):
    global name2t
    name2_cb.select_clear()
    name2t = name2.get()
    print(name2t)


def name3change(event):
    global name3t
    name3_cb.select_clear()
    name3t = name3.get()
    print(name3t)


name1 = tk.StringVar()
label = ttk.Label(namesfield, text='Techn. Person:').grid(row=1, column=0, sticky='E', pady=0)
name1_cb = ttk.Combobox(namesfield, textvariable=name1, value=poeple)
name1_cb.grid(row=1, column=1, sticky='Ew', pady=0)
name1_cb.bind('<Return>', name1change)
name1_cb.bind('<FocusOut>', name1change)
name1_cb.bind('<<ComboboxSelected>>', name1change)

# sample_entryfield.focus_set()

name2 = tk.StringVar()
label = ttk.Label(namesfield, text='Techn. Supervisor:').grid(row=2, column=0, sticky='E', pady=0)
name2_cb = ttk.Combobox(namesfield, textvariable=name2, value=sup, )
name2_cb.grid(row=2, column=1, sticky='Ew', pady=0)
name2_cb.set(sup[0])
name2_cb.bind('<Return>', name2change)
name2_cb.bind('<FocusOut>', name2change)
name2_cb.bind('<<ComboboxSelected>>', name2change)
name3 = tk.StringVar()
name3_cb = ttk.Label(namesfield, text='Med. Supervisor:')
name3_cb.grid(row=3, column=0, sticky='E')
name3_cb = ttk.Combobox(namesfield, textvariable=name3, value=doctors)
name3_cb.set(doctors[0])
name3_cb.grid(row=3, column=1, sticky='Ew')
name3_cb.bind('<Return>', name3change)
name3_cb.bind('<FocusOut>', name3change)
name3_cb.bind('<<ComboboxSelected>>', name3change)


# label = ttk.Label(text='Layers:').grid(row=4, column=0)
# layers_cb = ttk.Combobox(f1, textvariable='Layers:').grid(row=4, column=1)
def BatchPCR_entrychange(event):
    global BatchPCRt
    BatchPCRt = BatchPCR_entry.get()
    print(BatchPCRt)


BatchPCR_entry = tk.StringVar()
BatchPCR_entry = ttk.Label(namesfield, text='Batch PCR MM:').grid(row=4, column=0, sticky='E')
BatchPCR_entry = ttk.Entry(namesfield, textvariable='BatchN. PCR:')  # can be change to Entry or Combobox
BatchPCR_entry.grid(row=4, column=1, sticky='EW')
BatchPCR_entry.bind('<Return>', BatchPCR_entrychange)
BatchPCR_entry.bind('<FocusOut>', BatchPCR_entrychange)


def BatchRNA_entrychange(event):
    global BatchRNAt
    BatchRNAt = BatchRNA_entry.get()
    print(BatchRNAt)


BatchRNA_entry = tk.StringVar()
BatchRNA_entry = ttk.Label(namesfield, text='Batch RNA Ext.:').grid(row=5, column=0, sticky='E')
BatchRNA_entry = ttk.Entry(namesfield, textvariable='BatchN. RNA:')  # can be change to Entry or Combobox
BatchRNA_entry.grid(row=5, column=1, sticky='EW')
BatchRNA_entry.bind('<Return>', BatchRNA_entrychange)
BatchRNA_entry.bind('<FocusOut>', BatchRNA_entrychange)


# totalsamplelabel = ttk.Label(f1,text=f'Samples total: 0', font=('Sans 10 bold'))
# totalsamplelabel.grid(row=7, column=0,sticky='E')
# labelspucktest = ttk.Label(f1,text=f'Spucktest: {0}', font=('Sans', '10',))
# labelspucktest.grid(row=8, column=0, sticky='E')
# labelnonspucktest = ttk.Label(f1,text=f'Other: {0}', font=('Sans', '10',))
# labelnonspucktest.grid(row=9, column=0, sticky='E')
# labelduplicates = ttk.Label(f1,text=f'Duplicates: {0}', font=('Sans', '10',))
# labelduplicates.grid(row=10, column=0, sticky='E')
def layerselectorfunc(x):
    global layerSelected
    layerSelected = x
    UpdateLayer()
    print(f'layer {x + 1} is selected now')


foundLS = []


def Searchboxfunc(event=None):
    global layerSelected, newcell, foundLS
    msg = searchentry.get()
    cellfound = 0
    print(msg)
    if len(foundLS) > 0:
        for i in foundLS:
            print(i, msg)
            if i[2] != msg:
                foundLS = []
                Searchboxfunc()
            layerSelected = i[0]
            newcell = i[1]
            UpdateSelectedcell()
            UpdateLayer()
            try:
                foundLS.pop(0)
            except Exception:
                pass
            return
    print(foundLS)
    for count, layer in enumerate(listOfLayers):
        for count2, ID in enumerate(layer):
            if ID == msg:
                foundLS.append([count, count2, msg])
    if len(foundLS) == 0:
        tk.messagebox.showinfo(title='Not found', message=f'ID {msg} was not found')
    if len(foundLS) > 0:
        layerSelected = foundLS[0][0]
        newcell = foundLS[0][1]
        UpdateSelectedcell()
        UpdateLayer()
        foundLS.pop(0)
    searchentry.select_range(0, 'end')


def SearchboxClearfunc(event):
    if searchentry.get() == 'Search:':
        searchentry.delete(0, 'end')


def NoteshboxClearfunc(event):
    if notesentry.get("1.0", END) == 'Notes:\n':
        notesentry.delete("1.0", END)


def lineincellselectfunc(k):
    print('lineselected', k)
    global layerSelected, newcell, listOfLayers
    layerSelected = k
    a = newcell
    UpdateLayer()
    newcell = a
    UpdateSelectedcell()


def Genarate_buttons_fornewlayers():
    global sample_entryfieldold, sample_entryfield, lineselectorls, layerselectorls, listOfLayers
    for widget in cellview.winfo_children():
        widget.destroy()
    lineselectorls = []
    # lineselvariable=tk.StringVar()
    for count, value in enumerate(listOfLayers):
        lineselector = tk.Button(cellview, text=f'layer {count + 1}', height=1, width=27, font=("Courier", 12),
                                 wraplength=280,
                                 anchor="w",
                                 command=lambda k=count: lineincellselectfunc(k))
        lineselector.grid(row=1 + count, column=0, pady=2, padx=5)
        lineselectorls.append(lineselector)
    # layerselectorls = []
    # check_radio = tk.StringVar()
    # for count,value in enumerate(listOfLayers):
    #     layerselector = tk.Radiobutton(f1, text=f'layer {count+1}',variable=check_radio,value=f'{count}',font=('Sans 12 ')
    #                                    ,command=lambda k=count: layerselectorfunc(k))
    #     layerselector.grid(row=6+count, column=1)
    #     layerselectorls.append(layerselector)


Genarate_buttons_fornewlayers()

saerchentrystringvar = tk.StringVar(cellview, value='Search:')
searchentry = ttk.Entry(cellview, textvariable=saerchentrystringvar, width=24, font=('Sans 12 '))
# notesentrystringvar=tk.StringVar(cellview,value='Notes:')
# notesentry=ttk.Entry(cellview,textvariable=notesentrystringvar,width=24,font=('Sans 12 '))
searchentry.grid(row=0, column=0, sticky='we', pady=10, padx=10)
searchentry.bind('<Return>', Searchboxfunc)
searchentry.bind('<Tab>', Searchboxfunc)
searchentry.bind('<Button-1>', SearchboxClearfunc)


def export2Gentier():
    global listOfLayers, runname
    SaveFileOne(backup=1)
    gentiersample = ''
    for i in range(96):
        for count, sublist in enumerate(listOfLayers):
            if sublist[i] not in [None, '', 'empty', ' ']:
                sample = sublist[i]
                #Biothephy:
                # gentiersample += '{"ChInfo":[{"CH":1,"Concentration":0,"Index":16,"ProbeColor":"#FF00A0FF","TargetColor":"#FF00A0FF","TargetName":"NGENE","Unit":0},{"CH":2,"Concentration":0,"Index":32,"ProbeColor":"#FF00A000","TargetColor":"#FF00A000","TargetName":"ORF","Unit":0},null,{"CH":4,"Concentration":0,"Index":64,"ProbeColor":"#FFFF2020","TargetColor":"#FFFF2020","TargetName":"CONTROL","Unit":0},null,null],"Group":null,"Reference":0,"ReplicateGroup":"","SampleColor":"#FF4090B0","SampleId":"","SampleName":"' + str(
                #Allcheck:
                gentiersample += '{"ChInfo":[{"CH":1,"Concentration":0,"Index":16,"ProbeColor":"#FF00A0FF","TargetColor":"#FF00A0FF","TargetName":"ORF","Unit":0},{"CH":2,"Concentration":0,"Index":32,"ProbeColor":"#FF00A000","TargetColor":"#FF00A000","TargetName":"CONTROL","Unit":0},{"CH":3,"Concentration":0,"Index":48,"ProbeColor":"#FFFFA000","TargetColor":"#FFFFA000","TargetName":"NGENE","Unit":0},null,null,null],"Group":null,"Reference":0,"ReplicateGroup":"","SampleColor":"#FF2DB446","SampleId":"","SampleName":"' + str(
                    sample) + '","Type":"UNK","Uid":"","Well":' + str(i) + '},'
                break
    # for count,sample in enumerate(listOfLayers[0]):
    #     gentiersample+='{"ChInfo":[{"CH":1,"Concentration":0,"Index":16,"ProbeColor":"#FF00A0FF","TargetColor":"#FF00A0FF","TargetName":"NGENE","Unit":0},{"CH":2,"Concentration":0,"Index":32,"ProbeColor":"#FF00A000","TargetColor":"#FF00A000","TargetName":"ORF","Unit":0},null,{"CH":4,"Concentration":0,"Index":64,"ProbeColor":"#FFFF2020","TargetColor":"#FFFF2020","TargetName":"CONTROL","Unit":0},null,null],"Group":null,"Reference":0,"ReplicateGroup":"","SampleColor":"#FF4090B0","SampleId":"","SampleName":"'+str(sample)+'","Type":"UNK","Uid":"","Well":'+str(count)+'},'
    print('exporting')
    gentiersample = gentiersample[:-1]  # removes last comma XD
    gentiersamplecomplete = f'1[{gentiersample}]'
    print(gentiersamplecomplete)
    # with open(f'{runname}_g_export.tlps', 'w') as writer:
    folder2 = os.path.join(userpath + config['settings']['Pathfolder2'])
    # print(folder2+'\\'+runname+'_g_export.tlps')
    with open(folder2 + '\\' + runname + '_g_export.tlps', 'w') as writer:
        writer.write(gentiersamplecomplete)
    clipboard.copy(runname)  # puts runname in copybubuffer
    # gentierwindow = gw.getWindowsWithTitle('Skype')[0]
    gentierwindow = gw.getWindowsWithTitle('Real-time PCR system (user)')[0]
    print(gentierwindow._hWnd)
    win32gui.SetForegroundWindow(gentierwindow._hWnd)
    # gentierwindow.maximize()
    gentierwindow.restore()
    gentierwindow.activate()


exportbutton = tk.Button(controlsfield, text=f'Export to Gentier\n 0 samples', font='Helvetica 9 bold',
                         command=export2Gentier)
exportbutton.grid(row=0, column=0, sticky='EW', padx=2, pady=2)
ELISA_results = [[x] for x in range(96)]


def importform_A96():
    global ELISA_results
    ELISA_results = []
    dest_filename = os.path.join(userpath + config['settings']['ExportFromA96'] + '\\' + runname + '.csv')
    A96csvreadout = []
    with open(dest_filename, newline='\n') as csvfile:
        spamreader = csv.reader(csvfile, delimiter=';')
        for row in spamreader:
            A96csvreadout.append(row)
    print(f"{A96csvreadout=}")
    A96samplels = A96csvreadout[37:133]
    for row in A96samplels:
        print(f"{row=}")
    for i in range(96):
        # if listOfLayers[0][i] == '':
        #     continue
        meanOD = A96samplels[i][4]
        concentration = A96samplels[i][7]
        Result = ''
        Resultmanual = ''
        # if concentration >= config['settings']['SarsAB_cutoff']:
        #     Result = 'Positive'
        ELISA_results.append([i + 1, listCellnames[i], listOfLayers[0][i], meanOD, concentration, Result, Resultmanual])
    # pcr_resultsLS.append([wellnumber,well,ID,MeanOD,Concentration_(U/ml),Result,Resultmanual])
    print(f"results of a96redout function{ELISA_results=}")

def importFromGentier():
    if expvar.get() == 'ELISA':
        importform_A96()
        return
    # root.config(cursor="wait")
    # root.update()
    global pcr_resultsLS
    listCellnames = ['A1', 'A2', 'A3', 'A4', 'A5', 'A6', 'A7', 'A8', 'A9', 'A10', 'A11', 'A12', 'B1', 'B2', 'B3', 'B4',
                     'B5', 'B6', 'B7', 'B8', 'B9', 'B10', 'B11', 'B12', 'C1', 'C2', 'C3', 'C4', 'C5', 'C6', 'C7', 'C8',
                     'C9', 'C10', 'C11', 'C12', 'D1', 'D2', 'D3', 'D4', 'D5', 'D6', 'D7', 'D8', 'D9', 'D10', 'D11',
                     'D12', 'E1', 'E2', 'E3', 'E4', 'E5', 'E6', 'E7', 'E8', 'E9', 'E10', 'E11', 'E12', 'F1', 'F2', 'F3',
                     'F4', 'F5', 'F6', 'F7', 'F8', 'F9', 'F10', 'F11', 'F12', 'G1', 'G2', 'G3', 'G4', 'G5', 'G6', 'G7',
                     'G8', 'G9', 'G10', 'G11', 'G12', 'H1', 'H2', 'H3', 'H4', 'H5', 'H6', 'H7', 'H8', 'H9', 'H10',
                     'H11', 'H12']

    def importcurves():
        global dfNgene, dfORF, dfControl
        t = time.perf_counter()
        GentierToExcel = os.path.join(userpath + config['settings']['GentierToExcel'])
        dest_filename = f'{runname}[Abs Quant(Stage2_Step2)] (Stage2 - Step2) - Quantitation Amplification Results.xlsx'
        wb = load_workbook(filename=GentierToExcel + r'\\' + dest_filename)
        wsNgene = wb['Texas Red']
        wsORF = wb['FAM']
        wsControl = wb['HEX']
        dfNgene = pd.DataFrame(wsNgene.values)
        dfNgene.rename(columns=dfNgene.iloc[0], inplace=True)
        dfNgene.drop(dfNgene.index[0], inplace=True)
        dfORF = pd.DataFrame(wsORF.values)
        dfORF.rename(columns=dfORF.iloc[0], inplace=True)
        dfORF.drop(dfORF.index[0], inplace=True)
        dfControl = pd.DataFrame(wsControl.values)
        dfControl.rename(columns=dfControl.iloc[0], inplace=True)
        dfControl.drop(dfControl.index[0], inplace=True)
        elapsed_time = time.perf_counter() - t
        print(f"importcurves function {elapsed_time=}")

    if exporttime.get() != '':
        importcurves()
        return ()
    t = time.perf_counter()
    print('importing')
    GentierToExcel = os.path.join(userpath + config['settings']['GentierToExcel'])
    dest_filename = f'{runname}[Abs Quant(Stage2_Step2)] (Stage2 - Step2) - Quantitation Ct Results.xlsx'
    #try:
    wb = load_workbook(filename=GentierToExcel + r'\\' + dest_filename)
    elapsed_time = time.perf_counter() - t
    print(f"importFromGentier function loading workbook {elapsed_time=}")
    # except:
    #     print('Import from Gentier failed')
    #     tk.messagebox.showwarning(title='Import from Gentier failed', message='Gentier export file not found')
    ws = wb['0']
    data = ws.values
    columns = next(data)[0:]
    df = pd.DataFrame(data, columns=columns)
    lol = df.values.tolist()
    pcr_resultsLS = []
    nonmatchLS = [['Number,Well,Scanned ID, Gentier ID']]
    for i in range(int(len(lol) / 3)):
        # lol[i][5] is and ID
        i = i * 3 - 3
        well = lol[i][1]
        wellnumber = listCellnames.index(well) + 1
        # ID=lol[i][5] # get ID form the PCR file
        ID = ''
        for count, list in enumerate(listOfLayers):
            if listOfLayers[count][wellnumber - 1] not in [None, '', ' ']:
                ID += (str(listOfLayers[count][wellnumber - 1]) + '\n')  # use scanned ID's
        # print(f"{wellnumber=}")
        # print(f"{lol[i][5]=}")
        # print(f"{listOfLayers[0][wellnumber-1]=}")
        ID = ID[:-2]  # removes new line symbol
        if str(listOfLayers[0][wellnumber - 1]) != lol[i][5] and wellnumber <= 94:
            nonmatchLS.append((wellnumber, well, listOfLayers[0][wellnumber - 1], lol[i][5]))
        # for original kit where orf1	fam, N	hex, IC	cy5 :
        #NgeneCT = lol[i][6]
        #ORFCT = lol[i + 1][6]
        #ControlCT = lol[i + 2][6]
        # # for AstraF where fam	IC,hex	orf8,cy5	N,rox	orf1
        # NgeneCT=lol[i+2][6]
        # ORFCT=lol[i+1][6]
        # ControlCT=lol[i][6]
        # if ControlCT == 0:
        #for Allcheck:
        NgeneCT = lol[i+2][6]
        ORFCT = lol[i][6]
        ControlCT = lol[i + 1][6]
        Result = 'Repeat'
        # if 0<NgeneCT<37 and 0<ORFCT<37: # both genes need to be positive, non-zero, less that 37
        Ncutoff = int(config['settings']['NgeneCT'])
        ORFcutoff = int(config['settings']['ORFCT'])
        if config['settings']['GenesRequired'] == "2":
            if 0 < NgeneCT < Ncutoff and 0 < ORFCT < ORFcutoff:  # both genes need to be positive, non-zero, less that 37
                Result = 'Positive'
            elif ControlCT != 0:
                Result = 'Negative'
        else:
            if 0 < NgeneCT < Ncutoff or 0 < ORFCT < ORFcutoff:  # both genes need to be positive, non-zero, less that 37
                Result = 'Positive'
            elif ControlCT != 0:
                Result = 'Negative'
        Resultmanual = ''
        pcr_resultsLS.append([wellnumber, well, ID, NgeneCT, ORFCT, ControlCT, Result, Resultmanual])
    if len(nonmatchLS) > 1:
        tk.messagebox.showerror(title='ID mismatch detected',
                                message=f'Following IDs are mismatched:\n{nonmatchLS} \n '
                                        f'Mismathced File saved \n Make sure that you want to export this run')
        df = pd.DataFrame(nonmatchLS)
        Folder0 = os.path.join(userpath + config['settings']['Pathfolder0'] + r'\Mismatch Reports')
        df.to_csv(f'{Folder0}\{runname} Mismatch report.csv', index=False)
    pcr_resultsLS = sorted(pcr_resultsLS)
    SaveFileOne()
    print(f"{pcr_resultsLS=}")
    # LoadFileOne()
    Editrun.set(False)
    UpdateSelectedcell()
    elapsed_time = time.perf_counter() - t
    importcurves()
    root.config(cursor='')
    print(f"importFromGentier function {elapsed_time=}")


### Page 2 ####
global pcr_resultsLS
pcr_resultsLS = [[0]]


def onFrameConfigure(canvas):
    '''Reset the scroll region to encompass the inner frame'''
    canvas.configure(scrollregion=canvas.bbox("all"))


def Showpostable():
    DrawresulttablePage2()


def drawResolutionTable(event=0):
    scanforresolution()
    global dfResolution
    df = dfResolution.drop(['IDs', 'date', 'week', 'ispool', 'toResolve', 'daterun',
                            # 'NgeneCT','ORFCT','ControlCT',
                            'Result Auto'
                               , 'Manual Result'
                            # 'Control-CT'
                            ], axis=1)
    df.sort_values(by=['w.number'], inplace=True)
    df = df[['w.number', 'w.name', 'ID', 'NgeneCT', 'ORFCT', 'ControlCT', 'result', 'runwell', 'isResolved']]
    # print(df)
    # listofdicts=df.to_dict('records')
    listofdicts = df.values.tolist()
    print('########', listofdicts)
    # from pandastable import Table
    # pt = Table(f3, dataframe=dfResolution,editable=False,enable_menus=False,)
    # pt.show()
    columns = ('#1', '#2', '#3', '#4', '#5', '#6', '#7', '#8', '#9')
    tree = ttk.Treeview(f3, columns=columns, show='headings', height=38)  # int(52/len(listOfLayers)),) #50
    style.configure('Treeview',
                    # background='silver',
                    # foreground='black',
                    # fieldbackghroudm='silver',
                    font=("Courier", 12, 'bold'),
                    heigt=12,
                    rowheight=25,
                    # bordercolor='silver',
                    # lightcolor = "#ffc61e",
                    # bordercolor = "#ffc61e",
                    # darkcolor = "#ffc61e"
                    )
    style.map('Treeview', )
    # define headings
    tree.heading('#1', text='n.')
    tree.heading('#2', text='Well')
    tree.heading('#3', text='ID')
    tree.heading('#4', text='ORF-CT')
    tree.heading('#5', text='N-CT')
    tree.heading('#6', text='Control-CT')
    tree.heading('#7', text='Final Result')
    tree.heading('#8', text='Run well Pool')
    tree.heading('#9', text='Run well Repeat')
    tree.column('#1', width=60)
    tree.column('#2', width=60)
    tree.column('#4', width=60)
    tree.column('#5', width=60)
    tree.column('#6', width=60)
    tree.column('#3', width=300, anchor=W)
    tree.column('#8', width=250)
    tree.column('#9', width=250)

    # adding data to the treeview
    itteratorvar = 0
    for row in listofdicts:
        itteratorvar += 1
        if row[-1] == 'Not Found':
            tree.insert('', tk.END, values=row, tags='red')
        elif itteratorvar % 2 == 0:
            tree.insert('', tk.END, values=row, tags='even')
        else:
            tree.insert('', tk.END, values=row, tags='odd')
    tree.tag_configure('even', foreground='black', background='azure')
    tree.tag_configure('odd', foreground='black', background='gainsboro')
    tree.tag_configure('red', foreground='black', background='tomato')
    # bind the select event

    tree.grid(row=0, column=0, sticky='nsew', padx=(10, 0), pady=10, rowspan=45)

    changeresultbutton = tk.Button(f3, text='Refresh', command=lambda: [scanforresolution(), drawResolutionTable()])
    # changeresultbutton.bind('<Button-1>', command=lambda: [scanforresolution(),drawResolutionTable()])
    changeresultbutton.grid(row=3, column=3, padx=10, pady=10)

    # add a scrollbar
    scrollbar = ttk.Scrollbar(f3, orient=tk.VERTICAL, command=tree.yview)
    tree.configure(yscroll=scrollbar.set)
    scrollbar.grid(row=0, column=1, sticky='ns', pady=10, rowspan=45)
    return


def DrawresulttablePage2(wellnumber=0):
    global pcr_resultsLS, listOfLayers, Showposintable, positivetableLS, totalsamplesOnlyls, tree, listCellnames, dfNgene, canvas, \
        exportfinalresultbuttontime, conformfinalexport, exporttime

    def clearCanvascahrt():
        try:
            for item in canvas.get_tk_widget().find_all():
                canvas.get_tk_widget().delete(item)
        except:
            print('could not clear the chart')

    clearCanvascahrt()

    def UpdatePCRresultLS():
        for i in range(len(pcr_resultsLS)):
            actualwellnumber = pcr_resultsLS[i][0] - 1
            ids = ''
            for layer in listOfLayers:
                ids += f'{layer[actualwellnumber]}\n'
            ids = ids[:-2]
            pcr_resultsLS[i][2] = ids

    usedlayers = len([x for x in listOfLayers if not x[:94] == [''] * 94])
    if usedlayers == 0:
        usedlayers = 1
    if expvar.get() == 'ELISA':
        columns = ('#1', '#2', '#3', '#4', '#5', '#6', '#7')
        tree = ttk.Treeview(f2, columns=columns, show='headings', height=int(53 / usedlayers))  # in number of rows.
        resultsLS = ELISA_results
        tree.heading('#1', text='n.')
        tree.heading('#2', text='Well')
        tree.heading('#3', text='ID')
        tree.heading('#4', text='Mean OD')
        tree.heading('#5', text='Concentration (U/ml)')
        tree.heading('#6', text='Auto Result')
        tree.heading('#7', text='Manual Result')
        # tree.heading('#8', text='')
        tree.column('#1', width=60)
        tree.column('#2', width=60)
        tree.column('#3', width=250, anchor=W)
        tree.column('#4', width=150)
        tree.column('#5', width=190)
        tree.column('#6', width=190)
        tree.column('#7', width=190)
        # tree.column('#8', width=0)
    if expvar.get() == 'PCR':
        columns = ('#1', '#2', '#3', '#4', '#5', '#6', '#7', '#8',)
        tree = ttk.Treeview(f2, columns=columns, show='headings', height=int(53 / usedlayers))  # in number of rows.
        resultsLS = pcr_resultsLS
        UpdatePCRresultLS()
        namecol4 = 'N-CT'
        namecol5 = 'ORF-CT'
        namecol6 = 'Control-CT'
        tree.heading('#1', text='n.')
        tree.heading('#2', text='Well')
        tree.heading('#3', text='ID')
        tree.heading('#4', text=namecol4)
        tree.heading('#5', text=namecol5)
        tree.heading('#6', text=namecol6)
        tree.heading('#7', text='Auto Result')
        tree.heading('#8', text='Manual Result')
        tree.column('#1', width=60)
        tree.column('#2', width=60)
        tree.column('#3', width=250, anchor=W)
        tree.column('#4', width=150)
        tree.column('#5', width=150)
        tree.column('#7', width=190)
        tree.column('#8', width=190)
    print(f"{int(53/usedlayers)=}")
    style.configure('Treeview', rowheigh=25,
                    # fieldbackghroudm='silver',
                    font=("Courier", 12, 'bold'), heigt=12, rowheight=18 * usedlayers, )
    style.map('Treeview', )
    itteratorvar = 0
    positivetableLS = [x for x in resultsLS if (x[-1] in ['Positive', 'Unknown']
                                                or (x[-2] in ['Positive', 'Unknown'] and x[-1] not in ['Negative']))]
    # print('positivetableLS: ',positivetableLS)
    for samples in resultsLS:
        if (samples not in positivetableLS) and Showposintablevariable.get() == 1:
            continue
        else:
            itteratorvar += 1
            if itteratorvar % 2 == 0:
                tree.insert('', tk.END, values=samples, tags='even')
            else:
                tree.insert('', tk.END, values=samples, tags='odd')
    tree.tag_configure('even', foreground='black', background='azure')
    tree.tag_configure('odd', foreground='black', background='gainsboro')

    # bind the select event
    def buildcurve(selectedinTreewellnumber):
        global dfNgene, dfORF, dfControl, listOfLayers, positivetableLS, canvas
        # plt.close()
        # cell=listCellnames[selectedinTreewellnumber-1]
        targetwell = pcr_resultsLS[selectedinTreewellnumber - 1][0] - 1
        print(f"{targetwell=}")
        cell = listCellnames[targetwell]
        cellnumber = selectedinTreewellnumber - 1
        if Showposintablevariable.get() == 1:
            targetwell = positivetableLS[selectedinTreewellnumber - 1][0] - 1
            cell = listCellnames[targetwell]
            print('Showposintablevariable is TRUE : ' f"{cell=}")
        fig, ax = plt.subplots()
        dfControl[cell].plot(color='dodgerblue', label='Control')
        dfNgene[cell].plot(color='darkorange', label='N gene')
        dfORF[cell].plot(color='limegreen', label='ORF')
        plt.axhline(y=200, color='r', linestyle='--')
        plt.axvline(x=int(config['settings']['ORFCT']), color='r', linestyle='--')
        idstr = cell + ' ' + runname + " ID:" + (pcr_resultsLS[cellnumber][2]).replace('\n', ' ')
        print(f"{type(idstr)=}")
        plt.title(label=idstr,
                  fontsize=16,
                  color="k")
        ax.legend()
        plt.grid()
        canvas = FigureCanvasTkAgg(fig, master=f2)
        canvas.draw()
        toolbar = NavigationToolbar2Tk(canvas, f2, pack_toolbar=False)
        toolbar.update()
        canvas.draw()
        toolbar.grid(row=9, column=3, pady=0)
        canvas.get_tk_widget().grid(row=8, column=3, pady=0)
        # plt.show()

    def item_selected(event=None):  # on click
        global pcr_resultsLS, selectedinTreewellnumber, dfNgene, canvas
        for selected_item in tree.selection():
            selectedinTreewellnumber = (int(selected_item[1:], 16))
            print(f'{selectedinTreewellnumber=}')
        if not dfNgene.empty:
            thread = threading.Thread(None, buildcurve(selectedinTreewellnumber), None, (), {})
            thread.start()
        else:
            clearCanvascahrt()

    def Setmanualresult(event=0):
        thread = threading.Thread(None, SetmanualresultT, None, (), {})
        thread.start()

    def update_item(wellnumber):
        if expvar.get() == 'ELISA':
            resultsLS = ELISA_results
        if expvar.get() == 'PCR':
            resultsLS = pcr_resultsLS
        selected = tree.focus()
        tree.item(selected, values=(resultsLS[wellnumber - 1]))

    def SetmanualresultT(event=0):
        global pcr_resultsLS, ELISA_results, selectedinTreewellnumber, positivetableLS
        if expvar.get() == 'ELISA':
            resultsLS = ELISA_results
        if expvar.get() == 'PCR':
            resultsLS = pcr_resultsLS
        wellnumber = selectedinTreewellnumber
        if Showposintablevariable.get() == 1:
            targetwell = positivetableLS[selectedinTreewellnumber - 1][0]
            print(f'{targetwell=}')
            wellnumber = targetwell
        print(f'selected{selectedinTreewellnumber=}')
        previous_manual_result = resultsLS[wellnumber - 1][-1]
        if previous_manual_result in ['', 'Repeat']:
            resultsLS[wellnumber - 1][-1] = 'Positive'
        if previous_manual_result == 'Positive':
            resultsLS[wellnumber - 1][-1] = 'Negative'
        if previous_manual_result == 'Negative':
            resultsLS[wellnumber - 1][-1] = 'Unknown'
        if previous_manual_result == 'Unknown':
            resultsLS[wellnumber - 1][-1] = 'Repeat'

        if expvar.get() == 'ELISA':
            ELISA_results = resultsLS
        if expvar.get() == 'PCR':
            pcr_resultsLS = resultsLS
        update_item(wellnumber)
        # DrawresulttablePage2(wellnumber)
        tree.grid(row=0, column=0, sticky='nsew', padx=(10, 0), pady=10, rowspan=45)
        exportfinalresultbutton.configure(
            text=f'Export {ExportFinalCSV(countonly=1)} of {len(totalsamplesOnlyls)}\n samples to Final results\n To repeat:{len(totalsamplesOnlyls) - ExportFinalCSV(countonly=1)}',
            font='Helvetica 10 bold')
        print('number of exported samples=', ExportFinalCSV(countonly=1))
        SaveFileOne()

    tree.bind('<<TreeviewSelect>>', item_selected)
    tree.grid(row=0, column=0, sticky='nsew', padx=(10, 0), pady=10, rowspan=45)
    changeresultbutton = Button(f2, text='set manual result', command=Setmanualresult)
    # changeresultbutton.bind('<Button-1>',Setmanualresult)
    changeresultbutton.grid(row=1, column=3, padx=10, pady=10, sticky='w')
    exportfinalresultbutton = Button(f2,
                                     text=f'Export {ExportFinalCSV(countonly=1)} of {len(totalsamplesOnlyls)}\n samples to Final results\n To repeat:{len(totalsamplesOnlyls) - ExportFinalCSV(countonly=1)}',
                                     font='Helvetica 10 bold', command=ExportFinalCSV)
    # exportfinalresultbutton.bind('<Button-1>',ExportFinalCSV)
    exportfinalresultbutton.grid(row=5, column=3, sticky='w', padx=10, pady=10)
    exportfinalresultbuttontime = Label(f2, anchor="center", textvariable=exporttime, font=("Sans", 10,))
    exportfinalresultbuttontime.grid(row=6, column=3, sticky='w', padx=10, pady=10)
    conformfinalexport = tk.BooleanVar()
    conformfinalexport.set(False)
    conformfinalexportch = tk.Checkbutton()
    conformfinalexportch = ttk.Checkbutton(f2, text=f'Confirm {runname} is ready for export',
                                           variable=conformfinalexport,
                                           onvalue=1,
                                           offvalue=0,
                                           ).grid(row=3, column=3, sticky='w', padx=10, pady=10)
    # add a scrollbar
    scrollbar = ttk.Scrollbar(f2, orient=tk.VERTICAL, command=tree.yview)
    tree.configure(yscroll=scrollbar.set)
    scrollbar.grid(row=0, column=1, sticky='ns', pady=10, rowspan=45)
    raise_frame(f2)
    return


Showposintablevariable = tk.BooleanVar()
Showposintablevariable.set(False)
Checkbuttonp2 = ttk.Checkbutton(f2, text='Only Pos and Unk',
                                command=Showpostable,
                                variable=Showposintablevariable,
                                onvalue=1,
                                offvalue=0,
                                ).grid(row=2, column=3, sticky='w', padx=10, pady=10)
### Plate ###
listWithTitles = ['  ', '  1', '  2', '  3', '  4', '  5', '  6', '  7', '  8', '  9', '  10', '  11', '  12', '  ',
                  'A ', 'A 1', 'A 2', 'A 3', 'A 4', 'A 5', 'A 6', 'A 7', 'A 8', 'A 9', 'A 10', 'A 11', 'A 12', 'A ',
                  'B ', 'B 1', 'B 2', 'B 3', 'B 4', 'B 5', 'B 6', 'B 7', 'B 8', 'B 9', 'B 10', 'B 11', 'B 12', 'B ',
                  'C ', 'C 1', 'C 2', 'C 3', 'C 4', 'C 5', 'C 6', 'C 7', 'C 8', 'C 9', 'C 10', 'C 11', 'C 12', 'C ',
                  'D ', 'D 1', 'D 2', 'D 3', 'D 4', 'D 5', 'D 6', 'D 7', 'D 8', 'D 9', 'D 10', 'D 11', 'D 12', 'D ',
                  'E ', 'E 1', 'E 2', 'E 3', 'E 4', 'E 5', 'E 6', 'E 7', 'E 8', 'E 9', 'E 10', 'E 11', 'E 12', 'E ',
                  'F ', 'F 1', 'F 2', 'F 3', 'F 4', 'F 5', 'F 6', 'F 7', 'F 8', 'F 9', 'F 10', 'F 11', 'F 12', 'F ',
                  'G ', 'G 1', 'G 2', 'G 3', 'G 4', 'G 5', 'G 6', 'G 7', 'G 8', 'G 9', 'G 10', 'G 11', 'G 12', 'G ',
                  'H ', 'H 1', 'H 2', 'H 3', 'H 4', 'H 5', 'H 6', 'H 7', 'H 8', 'H 9', 'H 10', 'H 11', 'H 12', 'H ',
                  '  ', '  1', '  2', '  3', '  4', '  5', '  6', '  7', '  8', '  9', '  10', '  11', '  12', '  ']
listCellnames = ['A1', 'A2', 'A3', 'A4', 'A5', 'A6', 'A7', 'A8', 'A9', 'A10', 'A11', 'A12', 'B1', 'B2', 'B3', 'B4',
                 'B5', 'B6', 'B7', 'B8', 'B9', 'B10', 'B11', 'B12', 'C1', 'C2', 'C3', 'C4', 'C5', 'C6', 'C7', 'C8',
                 'C9', 'C10', 'C11', 'C12', 'D1', 'D2', 'D3', 'D4', 'D5', 'D6', 'D7', 'D8', 'D9', 'D10', 'D11', 'D12',
                 'E1', 'E2', 'E3', 'E4', 'E5', 'E6', 'E7', 'E8', 'E9', 'E10', 'E11', 'E12', 'F1', 'F2', 'F3', 'F4',
                 'F5', 'F6', 'F7', 'F8', 'F9', 'F10', 'F11', 'F12', 'G1', 'G2', 'G3', 'G4', 'G5', 'G6', 'G7', 'G8',
                 'G9', 'G10', 'G11', 'G12', 'H1', 'H2', 'H3', 'H4', 'H5', 'H6', 'H7', 'H8', 'H9', 'H10', 'H11', 'H12']
listCellnamesNumb = ['A1\n1', 'A2\n2', 'A3\n3', 'A4\n4', 'A5\n5', 'A6\n6', 'A7\n7', 'A8\n8', 'A9\n9', 'A10\n10',
                     'A11\n11', 'A12\n12', 'B1\n13', 'B2\n14', 'B3\n15', 'B4\n16', 'B5\n17', 'B6\n18', 'B7\n19',
                     'B8\n20', 'B9\n21', 'B10\n22', 'B11\n23', 'B12\n24', 'C1\n25', 'C2\n26', 'C3\n27', 'C4\n28',
                     'C5\n29', 'C6\n30', 'C7\n31', 'C8\n32', 'C9\n33', 'C10\n34', 'C11\n35', 'C12\n36', 'D1\n37',
                     'D2\n38', 'D3\n39', 'D4\n40', 'D5\n41', 'D6\n42', 'D7\n43', 'D8\n44', 'D9\n45', 'D10\n46',
                     'D11\n47', 'D12\n48', 'E1\n49', 'E2\n50', 'E3\n51', 'E4\n52', 'E5\n53', 'E6\n54', 'E7\n55',
                     'E8\n56', 'E9\n57', 'E10\n58', 'E11\n59', 'E12\n60', 'F1\n61', 'F2\n62', 'F3\n63', 'F4\n64',
                     'F5\n65', 'F6\n66', 'F7\n67', 'F8\n68', 'F9\n69', 'F10\n70', 'F11\n71', 'F12\n72', 'G1\n73',
                     'G2\n74', 'G3\n75', 'G4\n76', 'G5\n77', 'G6\n78', 'G7\n79', 'G8\n80', 'G9\n81', 'G10\n82',
                     'G11\n83', 'G12\n84', 'H1\n85', 'H2\n86', 'H3\n87', 'H4\n88', 'H5\n89', 'H6\n90', 'H7\n91',
                     'H8\n92', 'H9\n93', 'H10\n94', 'H11\n95', 'H12\n96']
listCellnamesNumb1 = ['A1 / 1', 'A2 / 2', 'A3 / 3', 'A4 / 4', 'A5 / 5', 'A6 / 6', 'A7 / 7', 'A8 / 8', 'A9 / 9',
                      'A10 / 10', 'A11 / 11', 'A12 / 12', 'B1 / 13', 'B2 / 14', 'B3 / 15', 'B4 / 16', 'B5 / 17',
                      'B6 / 18', 'B7 / 19', 'B8 / 20', 'B9 / 21', 'B10 / 22', 'B11 / 23', 'B12 / 24', 'C1 / 25',
                      'C2 / 26', 'C3 / 27', 'C4 / 28', 'C5 / 29', 'C6 / 30', 'C7 / 31', 'C8 / 32', 'C9 / 33',
                      'C10 / 34', 'C11 / 35', 'C12 / 36', 'D1 / 37', 'D2 / 38', 'D3 / 39', 'D4 / 40', 'D5 / 41',
                      'D6 / 42', 'D7 / 43', 'D8 / 44', 'D9 / 45', 'D10 / 46', 'D11 / 47', 'D12 / 48', 'E1 / 49',
                      'E2 / 50', 'E3 / 51', 'E4 / 52', 'E5 / 53', 'E6 / 54', 'E7 / 55', 'E8 / 56', 'E9 / 57',
                      'E10 / 58', 'E11 / 59', 'E12 / 60', 'F1 / 61', 'F2 / 62', 'F3 / 63', 'F4 / 64', 'F5 / 65',
                      'F6 / 66', 'F7 / 67', 'F8 / 68', 'F9 / 69', 'F10 / 70', 'F11 / 71', 'F12 / 72', 'G1 / 73',
                      'G2 / 74', 'G3 / 75', 'G4 / 76', 'G5 / 77', 'G6 / 78', 'G7 / 79', 'G8 / 80', 'G9 / 81',
                      'G10 / 82', 'G11 / 83', 'G12 / 84', 'H1 / 85', 'H2 / 86', 'H3 / 87', 'H4 / 88', 'H5 / 89',
                      'H6 / 90', 'H7 / 91', 'H8 / 92', 'H9 / 93', 'H10 / 94', 'H11 / 95', 'H12 / 96']
sampleindicatorls = [''] * 96
pos = 0
newcell = 0
oldcell = 0
buttonlist, celllabellist, celllabellist2 = [], [], []
editingoutoforder = 0
exporttime = StringVar()
exporttime.set(f'')
exportreportls = []


def cellbuttonFunc(k):
    if k in [94, 95] and expvar.get() == 'PCR':
        return
    global newcell, oldcell
    newcell = k
    UpdateSelectedcell()
    sample_entryfield.focus()


def sampleindicatorlsmaker():
    global dfResolution, listOfLayers, sampleindicatorls
    sampleindicatorls = [''] * 96
    for layer in listOfLayers:
        for count, ID in enumerate(layer):
            if ID not in [None, '', 'empty', 'POSITIVE', 'NEGATIVE', ' ']:
                if not dfResolution.empty:
                    # print(f"{dfResolution['ID'].tolist()}=")
                    if ID in dfResolution['ID'].tolist() and sum(listOfLayers, []).count(ID) > 1:
                        sampleindicatorls[count] += 'B'
                    elif ID in dfResolution['ID'].tolist():
                        sampleindicatorls[count] += 'R'
                    elif sum(listOfLayers, []).count(ID) > 1:
                        sampleindicatorls[count] += 'D'
                    else:
                        sampleindicatorls[count] += '*'
                elif sum(listOfLayers, []).count(ID) > 1:
                    sampleindicatorls[count] += 'D'
                else:
                    sampleindicatorls[count] += '*'
    # print(f'{sampleindicatorls=}')
    return sampleindicatorls


pos = 0
for i in range(8):
    for j in range(12):
        padtop = 0
        padbottom = 0
        padleft = 0
        padright = 0
        # if i%2==0:
        #     color='light cyan' #'gray92'
        # if i%2==1:
        #     color='light goldenrod'#'gray84'
        color = 'gray92'
        if i in [1, 3, 5]:
            padbottom = 2
        if i in [2, 4, 6]:
            padtop = 2
        if j in [2, 5, 8]:
            padright = 2
        if j in [3, 6, 9]:
            padleft = 2
        cellbutton = tk.Button(plateframe, bg=color, text=(listOfLayers[0][pos]), height=5, width=13,
                               font=("Courier", 11, 'bold'),
                               wraplength=120, anchor="center",
                               # relief=SUNKEN,
                               command=lambda k=pos: cellbuttonFunc(k))
        cellbutton.grid(row=i + 2, column=4 + j, padx=(padleft, padright), pady=(padtop, padbottom), )
        buttonlist.append(cellbutton)
        celllabel = tk.Label(plateframe, bg=color, anchor="nw", text=(listCellnamesNumb1[pos]), height=0, width=7,
                             font=("Sans", 10,), )
        celllabel.grid(row=i + 2, column=4 + j, sticky='nw', pady=(5, 0), padx=(5, 0))
        celllabel.bind('<Button-1>', lambda event, k=pos: buttonlist[k].invoke())
        celllabel2 = tk.Label(plateframe, bg=color, anchor="s", text=(sampleindicatorls[pos]), height=0, width=7,
                              font=("Sans", 10,), )
        celllabel2.grid(row=i + 2, column=4 + j, sticky='s', pady=(0, 5), padx=(0, 5))
        celllabel2.bind('<Button-1>', lambda event, k=pos: buttonlist[k].invoke())
        celllabellist.append(celllabel)
        celllabellist2.append(celllabel2)
        pos += 1


def Dailysummary():
    root.config(cursor="wait")
    root.update()
    scanforresolution()
    global dfDaily, exportreportls
    totaldaysample = dfDaily['ID'].tolist()
    totaldaysample = [x for x in totaldaysample if x not in ('NEGATIVE', 'POSITIVE')]
    totaldaysamplelocal = totaldaysample
    prefixdict2 = {y[0]: x for x, y in prefixdict.items()}
    prefixdict2original = prefixdict2.copy()
    samplelslocal = totaldaysample
    for key, value in prefixdict2original.items():
        prefixdict2[key] = len([x for x in samplelslocal if str(x).lower().startswith(str(value))])
        samplelslocal = [x for x in samplelslocal if not str(x).lower().startswith(str(value))]
    # print(f'{prefixdict2=}')
    Runsummarytext = tk.Text(prefixcountfield, width=38, height=14, font=("Courier 8"))
    summarystringLS = []
    summarystringLS.append(f'Date#{runname[0:8]}\n')
    summarystringLS.append(f'Total number of samples#{len(totaldaysamplelocal)}\n')
    for key, value in prefixdict2.items():
        # print(f"{key=}")
        # print(f"{value=}")
        if value != 0:
            summarystringLS.append(f'{key}#{value}\n')
    summarystringLS.append(f'.#.\n')
    summarystringLS.append(f'Other#{len(samplelslocal)}\n')

    summarystringLS.append(f'Repeated#{len(totalsamplesOnlyls) - len(set(totalsamplesOnlyls))}\n')
    # summarystringLS.append(f'Other#{ len([x for x in totalsamplesOnlyls if not str(x).lower().startswith(tuple(prefixdict2original.values()))])}\n')
    summarystringLSformated = []
    for line in summarystringLS:
        line = line.split('#')
        # print(f"{line=}")
        spacer = '.' * (38 - len(str(line[0])) - len(str(line[1])))
        line = f'{line[0]}{spacer}{line[1]}'
        summarystringLSformated.append(line)
    # print(f"{(summarystringLSformated)=}")
    Dailysummarytext = tk.Text(f4, width=38, height=44, font=("Courier 14"))
    Dailysummarytext.insert(1.0, ''.join(summarystringLSformated))
    Dailysummarytext.configure(state='disabled')
    Dailysummarytext.grid(row=0, column=0, sticky='news', pady=5, padx=5)
    for line in summarystringLS:
        line = line.split('#')
        # print(f"{line=}")
        spacer = '.' * (38 - len(str(line[0])) - len(str(line[1])))
        line = f'{line[0]}{spacer}{line[1]}'
        summarystringLSformated.append(line)
    exportreportstr = ''
    for line in natsorted(exportreportls):
        line = line.split('#')
        # print(f"{line=}")
        spacer = '.' * (60 - len(str(line[0])) - len(str(line[1])))
        line = f'{line[0]}{spacer}{line[1]}'
        exportreportstr += (line)
    exportreporttext = tk.Text(f4, width=60, height=44, font=("Courier 14"))
    exportreporttext.insert(1.0, ''.join(exportreportstr))
    exportreporttext.configure(state='disabled')
    exportreporttext.grid(row=0, column=1, sticky='news', pady=5, padx=5)
    raise_frame(f4)
    root.config(cursor="")


importbutton = tk.Button(controlsfield, text=f'Import results', font='Helvetica 10 bold',
                         command=lambda: [importFromGentier(), DrawresulttablePage2()]).grid(row=1, column=0,
                                                                                             sticky='EW', padx=2,
                                                                                             pady=2)
page2button = tk.Button(controlsfield, text=f'Results', font='Sans 10 bold',
                        command=lambda: [DrawresulttablePage2(), raise_frame(f2)]).grid(row=2, column=0, sticky='EW',
                                                                                        padx=2, pady=2)
page4Dailybutton = tk.Button(controlsfield, text=f'Summary', font='Sans 9 bold',
                             command=Dailysummary).grid(row=0, column=2, sticky='EW', padx=2, pady=2)
dfResolution = pd.DataFrame()


def resolutionbutton():
    global dfResolution
    page3button = tk.Button(controlsfield, text=f'repeats: empty', font='Sans 10 bold',
                            command=lambda: [raise_frame(f3), drawResolutionTable()])
    page3button.grid(row=2, column=1, sticky='EW', padx=2, pady=2)
    if (not dfResolution.empty) and ("Not Found" in dfResolution.isResolved.to_list()):
        page3button.config(
            text=f'repeats {len(dfResolution) - dfResolution.isResolved.value_counts()["Not Found"]}/ {len(dfResolution)}')


def experimentchange(event):
    global totalNumberofLayers
    showinfo(title='Exp', message=f'new experiment is {expvar.get()} {expTypeVar.get()}')
    if expvar.get() == 'ELISA':
        totalNumberofLayers = 1
    if expvar.get() == 'PCR':
        totalNumberofLayers = 6
    Createnewrun()
    # Newrun


saverunbutton = tk.Button(namesfield, text=f'Save', font='Sans 8 bold',
                          command=lambda: [SaveFileOne()]).grid(row=1, column=2, sticky='EW', padx=0, pady=0)
loadrunbutton = tk.Button(namesfield, text=f'Load', font='Sans 8 bold',
                          command=lambda: [LoadFileOne()]).grid(row=3, column=2, sticky='EW', padx=0, pady=0)
audioswitch = tk.BooleanVar()
audioswitch.set(config['settings']['speech'])
audioswitchbtn = ttk.Checkbutton(controlsfield, text='Announce well', variable=audioswitch, onvalue=1, offvalue=0,
                                 ).grid(row=16, column=0, sticky='e', padx=2, pady=2)
tk.Label(controlsfield, text="Experiment Type:").grid(row=17, column=0, padx=2)
experimenttypecb = ttk.Combobox(controlsfield, textvariable=expvar, values=['PCR', 'ELISA'], state="readonly", width=5)
experimenttypecb.grid(row=17, column=1, padx=2)
experimenttypecb.bind('<<ComboboxSelected>>', experimentchange)

experimenttypecb2 = ttk.Combobox(controlsfield, textvariable=expTypeVar, values=['SarsAB', 'VitaminD','IgE','othertest'], state="readonly", width=5)
experimenttypecb2.grid(row=17, column=2, padx=2)
experimenttypecb2.bind('<<ComboboxSelected>>', experimentchange)

audiolang = ttk.Combobox(controlsfield, text='Lang', values=['de', 'en'], state="readonly", width=5)
audiolang.grid(row=16, column=1, padx=2)
audiolang.set(config['settings']['lang'])

audiodelay = ttk.Combobox(controlsfield, text='Speechdelay', values=['0', '1', '2', '3', '4', '5'], width=5)
audiodelay.grid(row=16, column=2, padx=2)
audiodelay.set(config['settings']['Speechdelay'])

tk.Label(controlsfield, text="Layers:").grid(row=1, column=2, sticky='SW')
layerdepth = ttk.Combobox(controlsfield, text='layerdepth', values=['1', '2', '3', '4', '5', '6'], state="readonly",
                          width=5)
layerdepth.grid(row=2, column=2, padx=2, sticky='NW')
layerdepth.set('1')

notesentry = tk.Text(notesfield, width=38, height=7, font=('Sans 10 '))

notesentry.insert(1.0, 'Notes:')
notesentry.grid(row=0, column=0, sticky='news', pady=5, padx=5)
notesentry.bind('<Return>', )
notesentry.bind('<Button-1>', NoteshboxClearfunc)
ys = ttk.Scrollbar(notesfield, orient='vertical', command=notesentry.yview)
ys.grid(column=1, row=0, sticky='ns')
# def Changenumberoflayers(event): #not working
#     global totalNumberofLayers
#     msg=numberlayercom.get()
#     if msg != '':
#         AddLayers(totalNumberofLayers-int(msg))
#         totalNumberofLayers=int(msg)
#         print(totalNumberofLayers)
#         Genarate_buttons_fornewlayers()
# possiblenumber=list(range(1,11))
# numberlayerco=tk.StringVar()
# numberlayercom = ttk.Combobox(f1, textvariable=numberlayerco, value=possiblenumber)
# numberlayercom.grid(row=14, column=0)
# numberlayercom.bind('<<ComboboxSelected>>', Changenumberoflayers)
# numberlayercom.bind('<Return>', Changenumberoflayers)
# numberlayercom.bind('<FocusOut>', Changenumberoflayers)
label = tk.Label(entryframe, text=f'Loading :', font=('Sans', '30',), width=7).grid(row=1, column=0, sticky='E')
label = tk.Label(entryframe, text=f'Previous:', fg='gray70', font=('Sans', '30',), width=7).grid(row=0, column=0,
                                                                                                 sticky='E')

entrynumberlabelOld = tk.Label(entryframe, text=newcell, font=('Sans 30 bold'), bg="gray60", fg='gray70')
entrynumberlabelOld.grid(row=0, column=1, sticky='EW')
entrycoordlabelOld = tk.Label(entryframe, text=listCellnames[newcell], font=('Sans 30 bold'), bg="gray60", fg='gray70')
entrycoordlabelOld.grid(row=0, column=3, sticky='EW')

entrynumberlabel = tk.Label(entryframe, text=newcell, font=('Sans 30 bold'), bg="chocolate1")
entrynumberlabel.grid(row=1, column=1, sticky='EW')
entrycoordlabel = tk.Label(entryframe, text=listCellnames[newcell], font=('Sans 30 bold'), bg="chocolate1", )
entrycoordlabel.grid(row=1, column=3, sticky='EW')


# def WellLoadingcheck():
#     global newcell, layerSelected
#     layers=layerdepth.get()

def PlateEndCheck():
    global newcell, layerSelected, totalNumberofLayers
    if newcell == Usedwells:
        layerSelectedold = layerSelected
        if totalNumberofLayers > 1:
            layerSelected = (layerSelected + 1) % totalNumberofLayers
        newcell = 0
        UpdateLayer()
        if layerSelectedold == totalNumberofLayers - 1:
            tk.messagebox.showinfo(title='Last well', message=f'Well 94 of {totalNumberofLayers}th layer was loaded')
    if newcell == -1:
        newcell = Usedwells - 1
        layerSelectedold = layerSelected
        layerSelected = (layerSelected - 1) % totalNumberofLayers
        UpdateLayer()


def LayerEndCheck():
    global layerSelected


lastEntryTime = time.time()


def sample_entry(event):  # on enter
    global newcell, oldcell, lastEntryTime, layerSelected
    msg = sample_entryfield.get()
    for i in msg:
        if i.lower() in 'ä ü ß ö '.split() or not i.isalnum():
            tk.messagebox.showwarning(title='Strange symbol detected',
                                      message=f'Strange symbol "{i}" detected- check keyboard layout, capslock and the barcode')
    if expvar.get() == 'PCR':
        if msg in totalsamplesOnlyls and listOfLayers[layerSelected][newcell] != (msg):
            # ignore duplicates:
            sample_entryfield.delete(0, 'end')
            return
            # Error for duplicates:
            # tk.messagebox.showwarning(title='Duplicate', message='Duplicate scan warning')
        if msg.isdigit() and len(msg) != 13 and not msg.startswith("40000"):
            tk.messagebox.showwarning(title='Wrong length ID',
                                      message=f'Numberical ID are expected to be 13 digit long')
        elif (not msg.lower().startswith(("scp", 'priv', "40000"))) and len(msg) != 12 and len(msg)!=0:
            tk.messagebox.showwarning(title='Wrong length ID',
                                      message=f'Prefixed IDs should expected to be 12 symbols long')
    if time.time() - lastEntryTime < .5 and msg != '':
        print(f'{lastEntryTime},{time.time()}')
        tk.messagebox.showwarning(title='too fast',
                                  message=f'Fast consecutive scan, check barcode. \n 2 last scans rejected')
        newcell -= 1
        UpdateSelectedcell()
        return
    lastEntryTime = time.time()
    # listOfLayers[layerSelected][abs(newcell-1)]=('') #clears cell before entry - No :)
    listOfLayers[layerSelected][newcell] = (msg)  # changes record in the list
    sample_entryfield.delete(0, 'end')  # clean entry field
    ###
    print(f'{int(layerdepth.get())=}')
    if int(layerdepth.get()) == 1:
        if expvar.get() == 'ELISA':
            newcell = verticallist[verticallist.index(newcell) + 1]
        else:
            newcell += 1
    else:
        layerSelected += 1
        if layerSelected >= int(layerdepth.get()):
            newcell += 1
            layerSelected = 0
    UpdateLayer()
    # layerSelected+=1
    UpdateSelectedcell()
    PlateEndCheck()
    LayerEndCheck()
    SaveFileOne()


def SaveFileOne(event=0, backup=0):
    t = time.perf_counter()
    global runname, name1t, name2t, name3t, totalNumberofLayers, newcell, layerSelected, pcr_resultsLS, BatchPCRt, BatchRNAt, listOfLayers, todaysfilesLS, expvar
    savedHeadTitles = ['Runname', 'Techn. Person', 'Techn. Supervisor', 'Med. Supervisor', 'Batch PCR MM:',
                       'Batch RNA Ext.:', 'totalNumberofLayers', 'layerSelected', 'newcell', 'Notes', 'Final Export',
                       'Type']
    savedHead = [runname, name1t, name2t, name3t, BatchPCRt, BatchRNAt, totalNumberofLayers, layerSelected, newcell,
                 notesentry.get("1.0", END), exporttime.get(), expvar.get()]
    if expvar.get() == 'PCR':
        savedHeadTitles2 = ['w.number', 'w.name'] + [f'layer {i + 1}' for i in range(totalNumberofLayers)] + ["NgeneCT",
                                                                                                          "ORFCT",
                                                                                                          'ControlCT',
                                                                                                          'Result Auto',
                                                                                                          'Manual Result']
    if expvar.get() == 'ELISA':
        savedHeadTitles2 = ['w.number', 'w.name'] + [f'layer {i + 1}' for i in range(totalNumberofLayers)] + ["Mean OD",
                                                                                                              "Concentration (U/ml)",
                                                                                                              'Result Auto',
                                                                                                              'Manual Result']
    #
    wb = Workbook()
    ws0 = wb.create_sheet("Sheet", 0)
    try:
        ws0.append(savedHeadTitles)
        ws0.append(savedHead)
        ws0.append([''])
        ws0.append(savedHeadTitles2)
        for sampleid in range(96):
            cell_ids_ls = [sampleid + 1, listCellnames[sampleid]]
            for layer in range(totalNumberofLayers):
                cell_ids_ls.append(listOfLayers[layer][sampleid])
            pcrsavevalues = [x for x in pcr_resultsLS if x[0] == sampleid + 1]
            if len(pcrsavevalues) > 0 and expvar.get() == 'PCR':
                cell_ids_ls.extend(pcrsavevalues[0][3:])
            if expvar.get() == 'ELISA':
                cell_ids_ls = [sampleid + 1, listCellnames[sampleid]]
                cell_ids_ls.append(listOfLayers[layer][sampleid])
                if ELISA_results!=[]:
                    cell_ids_ls=ELISA_results[sampleid]
            ws0.append(cell_ids_ls)
        for i in range(3, 3 + totalNumberofLayers):
            ws0.column_dimensions[
                get_column_letter(i)].width = 15  # set size of ID wells 26 to fill full long barcrodes
        Folder1 = os.path.join(userpath + config['settings']['Pathfolder1'])
        if expvar.get() == 'ELISA':
            Folder1 = os.path.join(userpath + config['settings']['ElISAsaves'])
        if backup == 1:
            Folder1 = os.path.join(userpath + config['settings']['Pathfolder1'] + r'\backup')
            print('savingbackup', Folder1)
        wb.save(Folder1 + '\\' + runname + '.xlsx')

    except:
        tk.messagebox.showwarning(title='Error', message=sys.exc_info()[1])
        print(sys.exc_info())
    thread = threading.Thread(None, scanforresolution, daemon=True)
    thread.start()
    elapsed_time = time.perf_counter() - t
    print(f"SaveFileOne function {elapsed_time=}")
    todaysfilesLS = getfileLSfromdate(runname)


def LoadFileOne(event=0):
    t = time.perf_counter()
    # root.config(cursor="pirate")
    global runname, name1t, name2t, name3t, totalNumberofLayers, newcell, layerSelected, pcr_resultsLS, BatchPCRt, BatchRNAt, \
        listOfLayers, Editrun, userpath, todaysfilesLS, tree, dfNgene, exporttime,ELISA_results
    dfNgene = pd.DataFrame
    Folder1 = os.path.join(userpath + config['settings']['Pathfolder1'])
    if expvar.get() == 'ELISA':
        Folder1 = os.path.join(userpath + config['settings']['ElISAsaves'])
    wb = load_workbook(filename=(Folder1 + '\\' + runname + '.xlsx'))
    ws0 = wb.worksheets[0]
    for row in ws0.iter_rows(min_row=2, max_col=11, max_row=2, values_only=True):
        configValues = row
    loadHeadvalues = []
    for count, cell in enumerate(configValues):
        if cell == None:
            cell = ''
        loadHeadvalues.append(cell)
        print(f'{count=},{cell=}')

    [runname, name1t, name2t, name3t, BatchPCRt, BatchRNAt, totalNumberofLayersTemp, layerSelected, newcell, notes,
     exporttimeval] = loadHeadvalues
    id_entry.delete(0, tk.END)
    id_entry.insert(0, runname)
    name1_cb.delete(0, tk.END)
    name1_cb.insert(0, name1t)
    name2_cb.delete(0, tk.END)
    name2_cb.insert(0, name2t)
    name3_cb.delete(0, tk.END)
    name3_cb.insert(0, name3t)
    BatchPCR_entry.delete(0, tk.END)
    BatchPCR_entry.insert(0, BatchPCRt)
    BatchRNA_entry.delete(0, tk.END)
    BatchRNA_entry.insert(0, BatchRNAt)
    notesentry.delete('1.0', tk.END)
    notesentry.insert('1.0', notes)
    exporttime.set(exporttimeval)
    print(f'{name1t=}')
    # if totalNumberofLayersTemp>totalNumberofLayers:
    #     totalNumberofLayers=totalNumberofLayersTemp
    print(f'{loadHeadvalues=}')
    print('config output', runname, name1t, name2t, name3t, BatchPCRt, BatchRNAt, totalNumberofLayers, layerSelected,
          newcell)
    # AddLayers(totalNumberofLayers)
    listOfLayers = []
    pcr_resultsLS = []
    columnsls = []
    for col in ws0.iter_cols(min_row=5, max_row=96 + 4, min_col=0, max_col=(2 + totalNumberofLayers + 5),
                             values_only=True):
        layer = list(['' if x in [None, 'None'] else x for x in col])
        columnsls.append(layer)
        # print(f'{len(listOfLayers)=}')
    listOfLayers = columnsls[2:2 + totalNumberofLayers]
    if expvar.get() == 'ELISA':
        ELISA_results=columnsls[:]
        ELISA_results=[[row[i] for row in ELISA_results] for i in range(96)]
    pcr_resultsLS_col = columnsls[0:2 + totalNumberofLayers]
    pcr_resultsLS_col.extend(columnsls[2 + totalNumberofLayers:])
    # print(listOfLayers)
    # print('########################')
    pcr_resultsLS = [[r[col] for r in pcr_resultsLS_col] for col in range(len(pcr_resultsLS_col[0]))]
    for count, row in enumerate(pcr_resultsLS):
        ids = ''
        for i in range(2 + totalNumberofLayers):
            # if row[i + 2] in [None, '', 'empty', ' ']:
            #     break
            ids += f'{row[i + 2]}\n'
        ids = ids[:-2]
        pcr_resultsLS[count] = [row[0], row[1], ids, row[-5], row[-4], row[-3], row[-2], row[-1]]
    # pcr_resultsLS=[[j for j in] for i in pcr_resultsLS]
    # print(listOfLayers)
    # print(pcr_resultsLS_col)
    # print(pcr_resultsLS)
    try:
        if pcr_resultsLS[-1][-2] == '':
            print(f"{pcr_resultsLS=}")
            Editrun.set(True)
        elif pcr_resultsLS[-1][-2] != '':
            # print(f"{pcr_resultsLS=}")
            Editrun.set(False)
    except:
        print('runs PCR results is not empty')
    thread = threading.Thread(None, scanforresolution, daemon=True)
    thread.start()
    UpdateLayer()
    UpdateSelectedcell()
    root.title(f'{runname} -Comboa {version}')
    # print(f'{(time.time()-start)=}')
    sample_entryfield.focus_set()
    elapsed_time = time.perf_counter() - t
    print(f"LoadFileOne {elapsed_time=}")
    todaysfilesLS = getfileLSfromdate(runname)
    # root.config(cursor="")
    tree.destroy()
    return


def UploadToSQLDB(path):
    df = pd.read_csv(path, sep=';', encoding='utf8')
    df.drop(columns='Name', inplace=True)
    engine = create_engine(
        "mysql+pymysql://{user}:{pw}@/{db}".format(user="", pw="",
                                                                db=""))
    df.rename(columns={'UUID': 'SampleID', 'PCR result': 'Result'}, inplace=True)
    tablename = 'resultsonly'
    df[['TestDay', 'TestTime']] = df['Probeneingang'].str.split(' ', expand=True)
    df['TestTime'] = df['TestTime'].str.slice(stop=5)
    df['location'] = f"{config['settings']['Region2']}"
    df['DeliveryStatus'] = 'analyzed'
    df['ValidatedBy'] = 'NAME, MD'
    df.drop(columns='Probeneingang', inplace=True)
    df.drop(columns='Zentrum', inplace=True)
    try:
        df.to_sql(tablename, con=engine, if_exists='append', index=False, chunksize=1000)
        print('UploadToSQLDB')
    except:
        tk.messagebox.showerror(title='SQL upload Failed',
                                message=f'SQL upload Failed, \nCheck Internet concetion, or upload \nfile to the web dropbox manually')
def UploadToSQLDBElisa(path):
    string=''
    engine = create_engine(
        "mysql+pymysql://{user}:{pw}@/{db}".format(user="", pw="", db=""))
    with open(path) as csvfile:
        readCSV = csv.reader(csvfile, delimiter=';')
        next(readCSV)
        for row in readCSV:
            string = f"UPDATE DBS_Cards SET SarsAB_analysistime='{row[1]}', SarsAB='VALIDATED', SarsAB_result={row[3]} WHERE SampleID1='{row[0]}'; "
            print(f"{string=}")
            with engine.connect() as con:
                con.execute(string)
def ExportFinalCSV(event=0, countonly=0):
    # root.config(cursor="wait")
    current_time = datetime.now()
    dateandtime = current_time.strftime('%d.%m.%Y %H:%M:%S')
    global pcr_resultsLS, ELISA_results, listOfLayers, runname, prefixdict, userpath, exporttime
    prefixdict2 = {y[0]: x for x, y in prefixdict.items()}
    prefixdictfolder = {y: x[1] for y, x in prefixdict.items()}
    prefixdictname = {y: x[2] for y, x in prefixdict.items()}
    listCellnames = ['A1', 'A2', 'A3', 'A4', 'A5', 'A6', 'A7', 'A8', 'A9', 'A10', 'A11', 'A12', 'B1', 'B2', 'B3', 'B4',
                     'B5', 'B6', 'B7', 'B8', 'B9', 'B10', 'B11', 'B12', 'C1', 'C2', 'C3', 'C4', 'C5', 'C6', 'C7', 'C8',
                     'C9', 'C10', 'C11', 'C12', 'D1', 'D2', 'D3', 'D4', 'D5', 'D6', 'D7', 'D8', 'D9', 'D10', 'D11',
                     'D12', 'E1', 'E2', 'E3', 'E4', 'E5', 'E6', 'E7', 'E8', 'E9', 'E10', 'E11', 'E12', 'F1', 'F2', 'F3',
                     'F4', 'F5', 'F6', 'F7', 'F8', 'F9', 'F10', 'F11', 'F12', 'G1', 'G2', 'G3', 'G4', 'G5', 'G6', 'G7',
                     'G8', 'G9', 'G10', 'G11', 'G12', 'H1', 'H2', 'H3', 'H4', 'H5', 'H6', 'H7', 'H8', 'H9', 'H10',
                     'H11', 'H12']
    exportLS = []
    if expvar.get() == 'PCR':
        for wellresults in pcr_resultsLS:
            try:
                wellname = wellresults[1]  # A1
                wellnumber = listCellnames.index(wellname) + 1
                if wellresults[7] == '':
                    result = wellresults[6]  # manual result
                else:
                    result = wellresults[7]  # autoresult
                ##CT
                if wellresults[3] > 0 and wellresults[4] > 0:
                    # CTwell= round(min(wellresults[3:5]),2) #min
                    CTwell = round(sum(wellresults[3:5]) / 2, 2)  # avrg
                else:
                    CTwell = max(wellresults[3:5])
                if CTwell == 0 or result == 'Negative':
                    CTwell = ''
                if CTwell != '':
                    CTwell = str(CTwell).replace('.', ',')
                    # CTwell=str(CTwell)[0:2]+','+str(CTwell)[3:5]
                IDlist = []
                for layer in listOfLayers:
                    ID = layer[wellnumber - 1]  # ID of samples in the well, corresponding to PCR result
                    if (ID not in [None, '', 'POSITIVE', 'NEGATIVE', ' ']
                            and not result == 'Repeat'):
                        IDlist.append(ID)
                if len(IDlist) > 1 and result in ['Positive',
                                                  'Repeat']:  # count number of ids in the well, to skip if more than 1
                    IDlist = []  # clear export nothing from this well
                for id in IDlist:
                    lineinfinallist = (f'{id};{dateandtime};{result};{CTwell}')
                    exportLS.append(lineinfinallist)
            except:
                # print('nothing to export')
                continue
    if expvar.get() == 'ELISA':
        current_time = datetime.now()
        dateandtime = current_time.strftime('%Y-%m%d %H:%M:%S')
        for wellresults in ELISA_results:
            try:
                wellname = wellresults[1]  # A1
                wellnumber = listCellnames.index(wellname) + 1
                if wellresults[-1] == '':
                    result = wellresults[-2]  # manual result
                else:
                    result = wellresults[-1]  # autoresult
                concentration = wellresults[4]
                concentration = str(concentration).replace(',','.')
                ##CT
                IDlist = []
                for layer in listOfLayers:
                    ID = layer[wellnumber - 1]  # ID of samples in the well, corresponding to PCR result
                    if (ID not in [None, '', 'POSITIVE', 'NEGATIVE', ' ']
                            and not result == 'Repeat'):
                        IDlist.append(ID)
                if len(IDlist) > 1 and result in ['Positive',
                                                  'Repeat']:  # count number of ids in the well, to skip if more than 1
                    IDlist = []  # clear export nothing from this well
                for id in IDlist:
                    lineinfinallist = (f'{id};{dateandtime};MEASURED;{concentration}')
                    exportLS.append(lineinfinallist)
            except:
                # print('nothing to export')
                continue
    if countonly == 1:
        return (len(exportLS))
    if exporttime.get() != '':
        print(f"{exporttime.get()=}")
        tk.messagebox.showwarning(title='File was already exported',
                                  message='This file was already exported. \n Only one export is possible')
        return
    if conformfinalexport.get() == 0:
        print(f"{conformfinalexport.get()=}")
        tk.messagebox.showwarning(title='Confirmation required',
                                  message='Please check the results \n and confirm that they are ready for export')
        return
    if expvar.get() == 'PCR':
        uniqeclientdict = {}
        listoftotalprefixes = [k for k, v in prefixdictfolder.items()]
        for uniqeclient in set(prefixdictfolder.values()):
            folderpath = uniqeclient
            uniqeclientdict[uniqeclient] = [[k for k, v in prefixdictfolder.items() if v == uniqeclient]]
            # print(f'{uniqeclient=}')
            listofprefixesforthisclient = [k for k, v in prefixdictfolder.items() if v == uniqeclient]
            clientname = prefixdict[listofprefixesforthisclient[0]][2]
            # print(f'{listofprefixesforthisclient=}')
            # print(f'{folderpath=}')
            # print(f'{clientname=}')
            shortexportLS = [x for x in exportLS if (str(x).lower().split(';')[0]).startswith(
                tuple(listofprefixesforthisclient))]  # save per center
            exportLS = [x for x in exportLS if x not in shortexportLS]  # remove items from being double exported
            if uniqeclient == '06_Exp_F_Result_10_Sonstiges':
                shortexportLS = shortexportLS + [x for x in exportLS if not (str(x).lower().split(';')[0]).startswith(
                    tuple(
                        listoftotalprefixes))]  # if save in sonstige folder, saves all  probes with unexpected prefixes.
            shortexportLS.insert(0, 'UUID;Probeneingang;PCR result;CT;Zentrum;Name\n')
            # exportLS = [x for x in exportLS if not (str(x).lower().split(';')[0]).startswith(tuple(listofprefixesforthisclient))]  # remove already saved
            if len(shortexportLS) == 1:  # if nothing to export from client leave the loop
                continue
            if int(config['settings']['Cold_upload']) == 1:
                FolderExport = os.path.join(userpath + config['settings']['PathfolderExportCold'])
                exportfileaddress = FolderExport + '\\' + runname + '_' + clientname + '_FinalResult.csv'
            else:
                FolderExport = os.path.join(userpath + config['settings']['PathfolderExport'])
                exportfileaddress = FolderExport + '\\' + folderpath + '\\02_Final_validated_results' + '\\' + runname + '_' + clientname + '_FinalResult.csv'
            if os.path.isfile(exportfileaddress):
                print(f"{os.path.isfile(exportfileaddress)=} .. if True File already exists")
                tk.messagebox.showwarning(title='Final CSV alrady exists',
                                          message=f'Final CSV \n {exportfileaddress}\n alrady exists- Can not be replaced')
                return
            with open(exportfileaddress, "w", newline="\n") as f:
                for count, line in enumerate(shortexportLS):
                    if count > 0:
                        line = line + f';{clientname};\n'
                    f.write(line)
            UploadToSQLDB(exportfileaddress)
            print('CSV saved')
    if expvar.get() == 'ELISA':
        FolderExport = os.path.join(userpath + config['settings']['ELISAresults'])
        exportLS.insert(0,f'SampleID1;{expTypeVar.get()}_analysistime;{expTypeVar.get()};{expTypeVar.get()}_result;\n')
        exportfileaddress = FolderExport + '\\' + runname + '_FinalResult.csv'
        with open(exportfileaddress, "w", newline="\n") as f:
            for count, line in enumerate(exportLS):
                if count > 0:
                    line = line + f';\n'
                f.write(line)
        UploadToSQLDBElisa(exportfileaddress)
        print('CSV saved')
    current_time = datetime.now()
    dateandtime = current_time.strftime('%d.%m.%Y %H:%M:%S')
    exporttime.set(f'{runname} was exported at {dateandtime}')
    SaveFileOne()
    if int(config['settings']['Cold_upload']) == 0 and expvar.get() == 'PCR':
        SendEmail()
    # root.config(cursor="")


def SendEmail():  # make a gmail headless sender
    global runname, userpath
    emaildict = {i: config['Emails'][i].split(';') for i in config.options('Emails')}
    print(f'{emaildict=}')
    for folder, emails in emaildict.items():
        print(f'{emails[1:]}=')
        outlook = win32com.client.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = ';'.join(emails[1:])
        mail.Subject = f'{runname} Final result'
        mail.HTMLBody = '<h3>This is HTML Body</h3>'
        mail.Body = """ HTML text """
        FolderExport = os.path.join(userpath + config['settings']['PathfolderExport'])
        file = FolderExport + '06_Exp_F_Result_' + '\\02_Final_validated_results' + '\\' + runname + '_FinalResult.csv'
        print(f"{file=}")
        mail.Attachments.Add(file)
        mail.display()
        # mail.Send()


# sendEmailsbtn=Button(f2, text='Send email',command=SendEmail)
# sendEmailsbtn.grid(row=6, column=3, sticky='w',padx=10,pady=10)
def UpdateSelectedcell():  # and move to next well
    global newcell, oldcell, totalsamplesOnlyls, listOfLayers, pcr_resultsLS, foundLS, sampleindicatorls, Editrun, Usedwells
    totalsamplesls = []
    for i in listOfLayers:
        totalsamplesls = totalsamplesls + list(i)
    totalsamplesOnlyls = [i for i in totalsamplesls if i not in [None, '', 'empty', 'POSITIVE', 'NEGATIVE', ' ']]
    filledindexlist = [i for i in range(96 * totalNumberofLayers) if
                       totalsamplesls[i] not in [None, '', 'empty', 'POSITIVE', 'NEGATIVE', ' ']]
    totalsamples = len(filledindexlist)
    buttonlist[oldcell].config(bg='gray92')  # 3-4 times faster than turning all cells gray
    celllabellist[oldcell].config(bg='gray92')
    celllabellist2[oldcell].config(bg='gray92')
    sampleindicatorlsmaker()
    buttonlist[oldcell].config(relief=RAISED)
    # print(f"{len(pcr_resultsLS)=}")
    Usedwells = 96
    if expvar.get() == 'PCR':
        Usedwells = 94
    for count, button in enumerate(buttonlist):
        button.config(text=str(listOfLayers[layerSelected][count]))
        celllabellist2[count].config(text=sampleindicatorls[count])
        # if :
        if count < Usedwells:
            if len(pcr_resultsLS) > 1:  # if len(pcr_resultsLS[0])>1: ????
                # print(f"{pcr_resultsLS[count]=}")
                try:
                    currentline = [x for x in pcr_resultsLS if x[0] == count + 1][0]
                except:
                    print('current line is empty')
                # print(f"{currentline=}")
                resultsdifinitive = currentline[6].lower()
                if len(currentline[7]) > 0:
                    resultsdifinitive = currentline[7].lower()
                # print(f'{currentline=}')
                if resultsdifinitive == 'repeat':
                    button.config(bg='medium orchid')
                    celllabellist[count].config(bg='medium orchid')
                    celllabellist2[count].config(bg='medium orchid')
                elif resultsdifinitive == 'positive':
                    button.config(bg='tomato2')
                    celllabellist[count].config(bg='tomato2')
                    celllabellist2[count].config(bg='tomato2')
                else:
                    button.config(bg='gray92')
                    celllabellist[count].config(bg='gray92')
                    celllabellist2[count].config(bg='gray92')
            else:
                button.config(bg='gray92')
                celllabellist[count].config(bg='gray92')
                celllabellist2[count].config(bg='gray92')
        if expvar.get() == 'PCR':
            if count == 94:
                button.config(text='Negative')
            if count == 95:
                button.config(text='Positive')
    # if len(pcr_resultsLS[0]) < 2 or len(foundLS)>0:
    if Editrun.get() or len(foundLS) > 0:
        buttonlist[newcell].config(bg='green2')
        celllabellist[newcell].config(bg='green2')
        celllabellist2[newcell].config(bg='green2')
    if not Editrun.get():
        buttonlist[newcell].config(relief=SUNKEN)
    sample_entryfield.delete(0, 'end')  # clean entry field
    sample_entryfield.insert(0, str(listOfLayers[layerSelected][newcell]))
    sample_entryfield.select_range(0, 'end')
    entrynumberlabel.config(text=newcell + 1)
    entrycoordlabel.config(text=listCellnames[newcell])
    buttonlist[newcell].config(text=listOfLayers[layerSelected][newcell])
    entrynumberlabelOld.config(text=max(1, newcell))
    entrycoordlabelOld.config(text=listCellnames[max(0, newcell - 1)])
    sample_entryfieldold.config(text=listOfLayers[layerSelected][
        max(0, newcell - 1)])  # would need to include much logic to show well 94 from previous layer
    exportbutton.config(text=f'Export to Gentier\n {totalsamples} samples')
    # print(f'{len(listOfLayers)=}, {listOfLayers=}')
    oldcell = newcell
    for count, button in enumerate(lineselectorls):
        button.config(text=f'{count + 1}:{listOfLayers[count][newcell]}')
    print(f' {newcell=}')
    if 0 < newcell < 94 and audioswitch.get() == 1 and Editrun.get():
        # thread = threading.Thread(None, Playnewcell, None, (), {})
        # thread.setDaemon(True)
        thread = threading.Thread(None, Playnewcell, daemon=True)
        thread.start()
        # Playnewcell()
    resolutionbutton()
    Prefixcount()


def Prefixcount():
    global totalsamplesOnlyls, prefixdict
    prefixdict2 = {y[0]: x for x, y in prefixdict.items()}
    prefixdict2original = prefixdict2.copy()
    samplelslocal = totalsamplesOnlyls
    for key, value in prefixdict2original.items():
        prefixdict2[key] = len([x for x in samplelslocal if str(x).lower().startswith(str(value))])
        samplelslocal = [x for x in samplelslocal if not str(x).lower().startswith(str(value))]
    # print(f'{prefixdict2=}')
    fieldwidth = 35
    Runsummarytext = tk.Text(prefixcountfield, width=fieldwidth, height=14, font=("Courier 10"))
    summarystringLS = []
    summarystringLS.append(f'Total number of samples#{len(totalsamplesOnlyls)}\n')
    for key, value in prefixdict2.items():
        # print(f"{key=}")
        # print(f"{value=}")
        if value != 0:
            summarystringLS.append(f'{key}#{value}\n')
    summarystringLS.append(f'.#.\n')
    summarystringLS.append(f'Other#{len(samplelslocal)}\n')
    summarystringLS.append(f'Total number of samples#{len(totalsamplesOnlyls)}\n')
    summarystringLS.append(f'Duplicates#{len(totalsamplesOnlyls) - len(set(totalsamplesOnlyls))}\n')
    summarystringLSformated = []
    for line in summarystringLS[1:]:
        line = line.split('#')
        # print(f"{line=}")
        spacer = '.' * (fieldwidth - len(str(line[0])) - len(str(line[1])))
        line = f'{line[0]}{spacer}{line[1]}'
        summarystringLSformated.append(line)
    # print(f"{(summarystringLSformated)=}")
    Runsummarytext.insert(1.0, ''.join(summarystringLSformated))
    Runsummarytext.configure(state='disabled')
    Runsummarytext.grid(row=0, column=0, sticky='news', pady=5, padx=5)
    return


def Playnewcell():
    global lang
    lang = audiolang.get()
    # starttime=time.time()
    # print(f'{os.path.dirname(os.path.abspath(__file__))=}')
    time.sleep(float(audiodelay.get()))
    # print('text to speech duration',time.time()-starttime)
    # a=f'...{listCellnames[newcell]}'
    # tts = gTTS(text=a, lang=config['settings']['lang'], slow=False)
    # tts.save('text.mp3')
    # sound=AudioSegment.from_mp3('text.mp3')
    # pausesound= AudioSegment.silent(duration=float(config['settings']['Speechdelay'])*1000)
    # newsound=pausesound+sound
    # newsound.export('text.mp3',format='mp3')
    # lang=config['settings']['lang']
    audiopath = os.path.join(userpath + config['settings']['Pathfolder0'])
    playsound(f'{audiopath}//audio/{lang}/{newcell}.mp3', block=False)
    # print(sys.path[0])
    # print('text to speech duration',time.time()-starttime)
    # os.remove('text.mp3')


def UpdateLayer():
    global layerSelected, newcell, listOfLayers
    for count, button in enumerate(buttonlist):
        button.config(text=str(listOfLayers[layerSelected][count]))
    UpdateSelectedcell()
    # layerselectorls[layerSelected].select()
    for count, button in enumerate(lineselectorls):
        if layerSelected == count:
            button.config(bg='green2')
        else:
            button.config(bg='gray92')
    print(f'{layerSelected=}')


sample_entryfieldvar = tk.StringVar()
sample_entryfieldold = tk.Label(entryframe, fg='gray70', width=40, font=("Courier", 30, "bold"),
                                anchor='w')  # can be change to Entry or Combobox
sample_entryfieldold.grid(row=0, column=2, sticky='news')

sample_entryfieldvar = tk.StringVar()
sample_entryfield = ttk.Entry(entryframe, textvariable='entrynumber', width=40,
                              font=("Courier", 30, "bold"))  # can be change to Entry or Combobox
sample_entryfield.grid(row=1, column=2, sticky='news')
sample_entryfield.bind('<Return>', sample_entry)
Editrun = tk.BooleanVar(value=True)
Editcheckbttn = ttk.Checkbutton(controlsfield, text='Well indicator',
                                variable=Editrun,
                                onvalue=1, command=UpdateSelectedcell,
                                offvalue=0,
                                )
Editcheckbttn.grid(row=0, column=1, sticky='W', )


def selectsampleentrybox(event):
    sample_entryfield.focus_set()
    layerdepth.select_clear()


layerdepth.bind('<<ComboboxSelected>>', selectsampleentrybox)
name1_cb.bind('<<ComboboxSelected>>', selectsampleentrybox)
name2_cb.bind('<<ComboboxSelected>>', selectsampleentrybox)
name3_cb.bind('<<ComboboxSelected>>', selectsampleentrybox)


# sample_entryfield.bind('<FocusOut>', sample_entry)
# sample_entryfield.bind('<FocusOut>', lambda event, move=0: sample_entry(move))

def entryfieldPrevious(event):
    global newcell
    newcell -= 1
    print('going up')
    UpdateSelectedcell()
    PlateEndCheck()


def entryfieldNext(event):
    global newcell
    newcell += 1
    print('going down')
    PlateEndCheck()
    UpdateSelectedcell()


sample_entryfield.bind('<Up>', entryfieldPrevious)
# sample_entryfield.bind('<Left>',entryfieldPrevious) #left right better used for navigation in the text
sample_entryfield.bind('<Down>', entryfieldNext)
# sample_entryfield.bind('<Right>',entryfieldNext)


root.protocol("WM_DELETE_WINDOW", on_closing)
UpdateSelectedcell()
UpdateLayer()
raise_frame(f1)
thread = threading.Thread(None, scanforresolution, daemon=True)
thread.start()
sample_entryfield.focus_set()
root.mainloop()
